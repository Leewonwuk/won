"""v1.28 라이브 트레이딩 엔진.

상태 머신: IDLE → LEG_A_PENDING → LEG_B_PENDING → IDLE

DT ({COIN}USDT > {COIN}USDC):
  Leg A: {COIN}USDC 시장에서 {COIN} 매수 (지정가)
  Leg B: {COIN}USDT 시장에서 {COIN} 매도 (지정가, Leg A 체결 후 즉시)

DC ({COIN}USDC > {COIN}USDT):
  Leg A: {COIN}USDT 시장에서 {COIN} 매수 (지정가)
  Leg B: {COIN}USDC 시장에서 {COIN} 매도 (지정가, Leg A 체결 후 즉시)

타임아웃:
  Leg A 미체결 → 취소, IDLE 복귀
  Leg B 미체결 → 취소 후 시장가 비상 청산
"""
from __future__ import annotations

import csv
import logging
import threading
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta

_KST = timezone(timedelta(hours=9))
from enum import Enum
from pathlib import Path

from src.exchanges.binance_trading_client import BinanceTradingClient
from src.notifications.telegram_notifier import send_telegram, send_telegram_document
from src.v12.allocator_v2 import ActionV2, decide_v2
from src.v12.config_v2 import V12ConfigV2
from src.v12.global_lock import GlobalTradeLock
from src.v12.price_feed import PriceFeed
from src.v12.ws_order_client import BinanceTradingWsClient

# WS 이벤트가 없을 때 REST order status 폴백 최소 간격(초).
# WS 정상 시엔 거의 호출되지 않음.
_REST_ORDER_CHECK_INTERVAL = 0.5
# IOC miss 후 재시도 최소 간격(초)
_LEG_A_RETRY_COOLDOWN = 0.1

log = logging.getLogger(__name__)

# CSV 컬럼 순서
_CSV_FIELDS = [
    "거래시각(KST)", "거래번호", "방향",
    "신호프리미엄(%)",   # 신호 시점 raw_ask 기준 premium (threshold 판단 기준)
    "진입프리미엄(%)",   # Leg A 실체결가 기준 스프레드 (체결 후 재계산)
    "매수시장", "매수가격(USD)",
    "매도시장", "매도가격(USD)",
    "코인수량", "투입금액(USD)", "수취금액(USD)",
    "순수익(USD)", "순수익(원)",
    "수수료추정(USD)",
    "체결소요시간(초)",
    "거래후USDT", "거래후USDC",
    "총자산(USD)", "총자산(원)",
    "누적순수익(USD)", "누적순수익(원)",
    "잔여BNB",
    "비고",
]


class TradeState(str, Enum):
    IDLE = "IDLE"
    LEG_A_PENDING = "LEG_A_PENDING"
    LEG_B_PENDING = "LEG_B_PENDING"
    SPEC_PENDING  = "SPEC_PENDING"   # Speculative: 두 주문 동시 제출, 둘 다 모니터링 중


@dataclass
class V12LiveEngineV2:
    cfg: V12ConfigV2
    binance: BinanceTradingClient
    dry_run: bool = True
    poll_interval_sec: float = 0.5
    leg_timeout_sec: float = 3.0
    log_dir: str = "logs/v12"
    fx_krw: float = 1450.0   # KRW/USDT 환율 (.env ARB_FX_KRW_PER_USDT)
    # 텔레그램 운영 요약 + 거래 CSV 전송 주기 (초). 기본 1800=30분. live_v2_main 이 env 로 덮어씀.
    telegram_summary_sec: float = 1800.0
    # 실시간 WS 피드 (없으면 REST fallback)
    price_feed: PriceFeed | None = None
    # WS Order API 클라이언트 (없으면 REST fallback)
    ws_order_client: BinanceTradingWsClient | None = None
    # 멀티코인 통합 요약 모드: True 이면 개별 텔레그램 요약 억제 (main 이 통합 전송)
    suppress_hourly_telegram: bool = False
    # 통합 풀 운영 시 동시 진입 방지 락 (None = 분리 운영 모드)
    global_lock: GlobalTradeLock | None = None

    # 잔고
    usdt: float = field(default=0.0, init=False)
    usdc: float = field(default=0.0, init=False)
    bnb: float = field(default=0.0, init=False)

    # 상태 머신
    state: TradeState = field(default=TradeState.IDLE, init=False)
    current_action: ActionV2 = field(default=ActionV2.HOLD, init=False)
    leg_a_order_id: str = field(default="", init=False)
    leg_b_order_id: str = field(default="", init=False)
    leg_a_btc_qty: float = field(default=0.0, init=False)
    leg_a_notional: float = field(default=0.0, init=False)
    leg_a_price: float = field(default=0.0, init=False)
    leg_b_price: float = field(default=0.0, init=False)
    entry_premium: float = field(default=0.0, init=False)   # Leg A 실체결가 기준 (체결 후 재계산)
    signal_premium: float = field(default=0.0, init=False)  # 신호 시점 raw_ask 기준
    order_time: float = field(default=0.0, init=False)
    leg_a_fill_time: float = field(default=0.0, init=False)

    # 통계
    trade_count: int = field(default=0, init=False)
    total_pnl_usd: float = field(default=0.0, init=False)
    emergency_count: int = field(default=0, init=False)   # Leg B 타임아웃 시장가 청산
    abort_count: int = field(default=0, init=False)        # Leg B 진입 전 프리미엄 소멸 조기 청산
    timeout_count: int = field(default=0, init=False)      # Leg A 타임아웃 취소
    _log_seq: int = field(default=0, init=False)           # CSV 기록 순번 (정상+비상+조기청산 모두 포함)

    # 리밸런싱
    _last_rebalance_ts: float = field(default=0.0, init=False)

    # 30분 요약용
    _hour_trade_count: int = field(default=0, init=False)
    _hour_pnl: float = field(default=0.0, init=False)
    _hour_abort_count: int = field(default=0, init=False)
    _last_hourly_ts: float = field(default=0.0, init=False)
    _hour_max_dc: float = field(default=0.0, init=False)   # 구간 DC방향 최고 프리미엄
    _hour_max_dt: float = field(default=0.0, init=False)   # 구간 DT방향 최고 프리미엄
    _period_abort_count: int = field(default=0, init=False)
    _period_emergency_count: int = field(default=0, init=False)

    # CSV / 상태 파일 경로
    _csv_path: str = field(default="", init=False)
    _state_path: str = field(default="", init=False)
    # REST order status 폴백 마지막 호출 시각 (WS 연결 중엔 거의 사용 안 됨)
    _last_rest_check_ts: float = field(default=0.0, init=False)
    # Leg B 대기 중 가격 방어 체크 마지막 시각
    _last_price_check_ts: float = field(default=0.0, init=False)
    # Event-driven 루프용 wake 이벤트
    _wake_event: threading.Event = field(default_factory=threading.Event, init=False)
    # IOC miss 통계
    ioc_miss_count: int = field(default=0, init=False)
    # Leg A IOC miss 후 재시도 쿨다운
    _last_leg_a_attempt_ts: float = field(default=0.0, init=False)
    # IOC miss 전용 요약 (20분 단위)
    _last_ioc_summary_ts: float = field(default=0.0, init=False)
    _ioc_summary_sec: float = field(default=3600.0, init=False)  # 1시간
    _period_ioc_miss: int = field(default=0, init=False)         # 구간 miss 카운트
    # IOC miss 디버깅 로그 (Excel 출력용)
    _leg_a_raw_ask: float = field(default=0.0, init=False)       # Leg A 시그널 시점 원본 ask
    _leg_a_submit_ts: float = field(default=0.0, init=False)     # Leg A 주문 제출 시각
    _leg_a_via_ws: bool = field(default=False, init=False)       # WS Order API 사용 여부
    _ioc_miss_rows: list = field(default_factory=list, init=False)  # 누적 miss 행
    _ioc_miss_xlsx_path: str = field(default="", init=False)
    # 통합 이벤트 로그 (IOC miss + 조기청산 + 타임아웃)
    _event_rows: list = field(default_factory=list, init=False)
    _event_xlsx_path: str = field(default="", init=False)
    # 정기 잔고 동기화
    _last_periodic_sync_ts: float = field(default=0.0, init=False)
    # Speculative Leg B 상태 추적
    _spec_leg_a_filled: bool = field(default=False, init=False)
    _spec_leg_b_filled: bool = field(default=False, init=False)
    _spec_received: float = field(default=0.0, init=False)
    # Leg A 체결 후 Leg B 가격 재설정 여부 (틱마다 중복 실행 방지)
    _spec_leg_b_repriced: bool = field(default=False, init=False)
    # 최초 잔고 동기화 여부 (수익 복리 반영을 위한 초기/주기 분리)
    _initial_sync_done: bool = field(default=False, init=False)

    def __post_init__(self) -> None:
        Path(self.log_dir).mkdir(parents=True, exist_ok=True)
        tag = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
        sym = self.cfg.symbol.lower()
        self._csv_path = f"{self.log_dir}/trades_v2_{sym}_{tag}.csv"
        self._state_path = f"{self.log_dir}/engine_state_{sym}.json"
        self._ioc_miss_xlsx_path = f"{self.log_dir}/ioc_miss_debug_{sym}_{tag}.xlsx"
        self._event_xlsx_path = f"{self.log_dir}/event_debug_{sym}_{tag}.xlsx"
        # 헤더 작성
        with open(self._csv_path, "w", encoding="utf-8-sig", newline="") as f:
            csv.DictWriter(f, fieldnames=_CSV_FIELDS).writeheader()
        log.info(f"거래 CSV: {self._csv_path}")
        self._last_hourly_ts = time.time()
        self._load_state()

    # ── 유틸 ──────────────────────────────────────────────────────────────

    def _now_str(self) -> str:
        return datetime.now(tz=_KST).strftime("%Y-%m-%d %H:%M:%S")

    def _usd_krw(self, usd: float) -> str:
        """$12.34 → '12.34 USDT (약 17,893원)' 형식."""
        krw = usd * self.fx_krw
        return f"{usd:+.4f} USDT (약 {krw:+,.0f}원)"

    def _eq_krw(self, usd: float) -> str:
        """총자산용: '1,100.17 USDT (약 1,595,247원)' 형식."""
        krw = usd * self.fx_krw
        return f"{usd:,.2f} USDT (약 {krw:,.0f}원)"

    def _notify(self, msg: str, *, telegram_parse_mode: str | None = "HTML") -> None:
        log.info(msg)
        send_telegram(msg, parse_mode=telegram_parse_mode)

    def _send_csv(self, caption: str = "") -> None:
        """현재 거래 CSV 파일을 텔레그램으로 전송."""
        rows = max(0, sum(1 for _ in open(self._csv_path, encoding="utf-8-sig")) - 1)
        if rows == 0:
            return
        send_telegram_document(
            self._csv_path,
            caption or f"📎 거래기록 CSV ({rows}건)",
        )

    def _eff_fee(self) -> float:
        return self.cfg.fee_maker_rate if self.cfg.use_maker_fees else self.cfg.fee_rate

    def sync_balances(self) -> None:
        balances = self.binance.get_spot_balances()
        actual_usdt = balances.get("USDT", 0.0)
        actual_usdc = balances.get("USDC", 0.0)
        self.bnb = balances.get("BNB", 0.0)
        # 멀티코인 동시 운용 시 자본 충돌 방지:
        # - GlobalLock 통합 풀 모드: 실제 잔고를 initial 상한선으로 클램프.
        #   min(actual, tracked) 방식은 잔고가 한 방향으로만 줄어 DT 거래로
        #   USDT가 회복되어도 엔진이 이를 인식 못해 잔고 오인식 및 거짓
        #   리밸런싱이 발생하므로, GlobalLock 모드에서는 actual 기반으로 동기화.
        # - 분리 운영 모드(GlobalLock 없음): 기존 min(actual, tracked) 유지.
        if self.global_lock is not None:
            # 통합 풀 모드: 실제 잔고를 그대로 반영.
            # GlobalLock으로 1개 엔진만 거래하므로 이중계산 없음.
            # initial 캡 제거 → DC 수익으로 USDC가 initial 초과해도 올바르게 추적.
            self.usdt = actual_usdt
            self.usdc = actual_usdc
        elif not self._initial_sync_done:
            self.usdt = min(actual_usdt, self.cfg.initial_usdt)
            self.usdc = min(actual_usdc, self.cfg.initial_usdc)
        else:
            self.usdt = min(actual_usdt, self.usdt)
            self.usdc = min(actual_usdc, self.usdc)
        self._initial_sync_done = True
        log.info(
            f"[{self._now_str()}] 잔고 동기화: "
            f"USDT={self.usdt:.2f}(실제={actual_usdt:.2f})  "
            f"USDC={self.usdc:.2f}(실제={actual_usdc:.2f})  BNB={self.bnb:.6f}"
        )

    def _get_mids(self) -> tuple[float, float]:
        # WS 캐시 우선 (0ms latency)
        if self.price_feed is not None:
            mid_ut = self.price_feed.get_mid(self.cfg.binance_symbol_usdt)
            mid_uc = self.price_feed.get_mid(self.cfg.binance_symbol_usdc)
            if mid_ut is not None and mid_uc is not None:
                return mid_ut, mid_uc
            log.debug("[Engine] WS 캐시 미스 → REST fallback")
        # REST fallback
        t_ut = self.binance.get_book_ticker(self.cfg.binance_symbol_usdt)
        t_uc = self.binance.get_book_ticker(self.cfg.binance_symbol_usdc)
        return (
            (float(t_ut["bidPrice"]) + float(t_ut["askPrice"])) / 2.0,
            (float(t_uc["bidPrice"]) + float(t_uc["askPrice"])) / 2.0,
        )

    def _get_bid_ask(self, symbol: str):
        """(bid, ask) 쌍 반환 — WS 캐시 우선, REST fallback."""
        if self.price_feed is not None:
            ba = self.price_feed.get_bid_ask(symbol)
            if ba is not None:
                return ba
        t = self.binance.get_book_ticker(symbol)
        return float(t["bidPrice"]), float(t["askPrice"])

    def _get_best_ask(self, symbol: str) -> float:
        """Leg A IOC 매수용 best ask."""
        if self.price_feed is not None:
            ba = self.price_feed.get_bid_ask(symbol)
            if ba is not None:
                return ba[1]
        t = self.binance.get_book_ticker(symbol)
        return float(t["askPrice"])

    def _get_best_bid(self, symbol: str) -> float:
        """Leg B GTC 매도용 best bid."""
        if self.price_feed is not None:
            ba = self.price_feed.get_bid_ask(symbol)
            if ba is not None:
                return ba[0]
        t = self.binance.get_book_ticker(symbol)
        return float(t["bidPrice"])

    def _cancel_order_safe(self, symbol: str, order_id: str) -> None:
        """주문 취소 (실패해도 무시). WS 우선, REST fallback."""
        if not order_id or self.dry_run:
            return
        try:
            if self.ws_order_client and self.ws_order_client.is_ready():
                self.ws_order_client.cancel_order(symbol, order_id)
            else:
                self.binance.cancel_order(symbol, order_id)
        except Exception as e:
            log.warning(f"주문 취소 실패 (무시): {e}")

    def _premium(self, mid_ut: float, mid_uc: float) -> float:
        return (mid_ut - mid_uc) / mid_uc if mid_uc > 0 else 0.0

    def _on_price_update(self, sym: str, bid: float, ask: float) -> None:
        """WS bookTicker 콜백 (WS asyncio 스레드에서 호출). 빠른 경로 유지."""
        if self.state != TradeState.IDLE:
            return
        if time.time() - self._last_leg_a_attempt_ts < _LEG_A_RETRY_COOLDOWN:
            return
        # 실행가능 스프레드 계산 (bid/ask 기반 — mid 대비 더 정확한 진입 감지)
        feed = self.price_feed
        if feed is None:
            return
        sym_ut = self.cfg.binance_symbol_usdt
        sym_uc = self.cfg.binance_symbol_usdc
        if sym == sym_ut:
            # DC 감지: 매수 ask_USDT, 매도 bid_USDC
            other_ba = feed.get_bid_ask(sym_uc)
            if other_ba is None:
                return
            exec_premium = (other_ba[0] - ask) / ask  # (bid_USDC - ask_USDT) / ask_USDT
            threshold = self.cfg.dc_entry_threshold_rate
        elif sym == sym_uc:
            # DT 감지: 매수 ask_USDC, 매도 bid_USDT
            other_ba = feed.get_bid_ask(sym_ut)
            if other_ba is None:
                return
            exec_premium = (other_ba[0] - ask) / ask  # (bid_USDT - ask_USDC) / ask_USDC
            threshold = self.cfg.dt_entry_threshold_rate
        else:
            return
        if exec_premium >= threshold:
            self._wake_event.set()

    def _on_fill(self, order_id: int, ev: dict) -> None:
        """WS 체결/취소 이벤트 콜백 (WS asyncio 스레드에서 호출)."""
        self._wake_event.set()

    # ── CSV 로그 ──────────────────────────────────────────────────────────

    def _log_trade(
        self,
        action: str,
        leg_a_symbol: str,
        leg_b_symbol: str,
        leg_b_price: float,
        received: float,
        pnl: float,
        elapsed_sec: float,
        note: str = "",
    ) -> None:
        eff_fee = self._eff_fee()
        is_emergency = "EMERGENCY" in note or "STOP_LOSS" in note
        # Leg A는 IOC = taker 수수료
        # 정상거래 Leg B는 지정가 = maker 수수료
        # 비상청산/스탑로스 Leg B는 시장가 = taker 수수료
        if received > 0 and not is_emergency:
            fee_est = self.leg_a_notional * self.cfg.fee_rate + received * eff_fee
        else:
            # 비상청산: received는 시장가 수취금액, emergency_taker_fee_rate로 재계산
            # received=0인 경우(극히 드묾): 청산가 ≈ 매수가로 근사
            leg_b_notional = received if received > 0 else self.leg_a_notional
            fee_est = self.leg_a_notional * self.cfg.fee_rate + leg_b_notional * self.cfg.emergency_taker_fee_rate
        self._log_seq += 1
        equity = self.usdt + self.usdc
        row = {
            "거래시각(KST)": self._now_str(),
            "거래번호": self._log_seq,
            "방향": action,
            "신호프리미엄(%)": f"{self.signal_premium * 100:.4f}",
            "진입프리미엄(%)": f"{self.entry_premium * 100:.4f}",
            "매수시장": leg_a_symbol,
            "매수가격(USD)": f"{self.leg_a_price:.2f}",
            "매도시장": leg_b_symbol,
            "매도가격(USD)": f"{leg_b_price:.2f}",
            "코인수량": f"{self.leg_a_btc_qty:.8f}",
            "투입금액(USD)": f"{self.leg_a_notional:.4f}",
            "수취금액(USD)": f"{received:.4f}",
            "순수익(USD)": f"{pnl:.4f}",
            "순수익(원)": f"{pnl * self.fx_krw:.0f}",
            "수수료추정(USD)": f"{fee_est:.4f}",
            "체결소요시간(초)": f"{elapsed_sec:.1f}",
            "거래후USDT": f"{self.usdt:.4f}",
            "거래후USDC": f"{self.usdc:.4f}",
            "총자산(USD)": f"{equity:.4f}",
            "총자산(원)": f"{equity * self.fx_krw:.0f}",
            "누적순수익(USD)": f"{self.total_pnl_usd:.4f}",
            "누적순수익(원)": f"{self.total_pnl_usd * self.fx_krw:.0f}",
            "잔여BNB": f"{self.bnb:.6f}",
            "비고": note,
        }
        with open(self._csv_path, "a", encoding="utf-8-sig", newline="") as f:
            csv.DictWriter(f, fieldnames=_CSV_FIELDS).writerow(row)

    # ── 주기 운영 요약 (텔레그램 + CSV, 기본 30분) ─────────────────────────

    def pop_ioc_miss_rows(self) -> list:
        """누적된 IOC miss rows를 꺼내고 초기화 (통합 Excel용)."""
        rows = list(self._ioc_miss_rows)
        self._ioc_miss_rows.clear()
        return rows

    def pop_event_rows(self) -> list:
        """누적된 event rows를 꺼내고 초기화 (통합 Excel용)."""
        rows = list(self._event_rows)
        self._event_rows.clear()
        return rows

    def get_period_stats(self) -> dict:
        """현재 구간 통계를 반환 (통합 요약용). 카운터는 리셋하지 않음."""
        # Leg A는 IOC(taker) → fee_rate, Leg B는 GTC(maker) → _eff_fee()
        fee_stack = self.cfg.fee_rate + self._eff_fee() + 2.0 * self.cfg.slippage_rate
        return {
            "symbol":        self.cfg.symbol,
            "hour_trades":   self._hour_trade_count,
            "hour_pnl":      self._hour_pnl,
            "hour_abort":    self._hour_abort_count,
            "hour_max_dc":   self._hour_max_dc,
            "hour_max_dt":   self._hour_max_dt,
            "period_abort":      self._period_abort_count,
            "period_emergency":  self._period_emergency_count,
            "fee_stack":     fee_stack,
            "dt_threshold":  self.cfg.dt_entry_threshold_rate,
            "dc_threshold":  self.cfg.dc_entry_threshold_rate,
            "total_trades":  self.trade_count,
            "total_pnl":     self.total_pnl_usd,
            "usdt":          self.usdt,
            "usdc":          self.usdc,
            "bnb":           self.bnb,
            "abort_count":   self.abort_count,
            "emergency_count": self.emergency_count,
            "timeout_count": self.timeout_count,
            "ioc_miss_count": self.ioc_miss_count,
            "period_ioc_miss": self._period_ioc_miss,
        }

    def reset_period_stats(self) -> None:
        """구간 통계 리셋 (통합 요약 전송 후 main 이 호출)."""
        self._hour_trade_count = 0
        self._hour_pnl = 0.0
        self._hour_abort_count = 0
        self._hour_max_dc = 0.0
        self._hour_max_dt = 0.0
        self._period_ioc_miss = 0
        self._period_abort_count = 0
        self._period_emergency_count = 0
        self._last_hourly_ts = time.time()

    def _check_hourly(self) -> None:
        if time.time() - self._last_hourly_ts < self.telegram_summary_sec:
            return
        # 멀티코인 통합 모드: main 코디네이터가 텍스트 요약 + 통합 CSV 전송 담당
        # BUG-26: reset_period_stats()를 여기서 호출하면 main이 읽기 전에 카운터가 0이 됨.
        # → _last_hourly_ts만 갱신하고 리셋은 main(_combined_summary_loop)에 위임.
        if self.suppress_hourly_telegram:
            self._last_hourly_ts = time.time()
            return
        equity = self.usdt + self.usdc
        fee_stack_threshold = self.cfg.fee_rate + self._eff_fee() + 2.0 * self.cfg.slippage_rate
        dt_need = max(self.cfg.dt_entry_threshold_rate, fee_stack_threshold)
        dc_need = max(self.cfg.dc_entry_threshold_rate, fee_stack_threshold)
        span_min = max(1, int(round(self.telegram_summary_sec / 60.0)))
        self._notify(
            f"⏰ <b>v1.28 운영 요약 [{self.cfg.symbol}] ({span_min}분)</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"🔄 이번 구간 거래: {self._hour_trade_count}회\n"
            f"💵 이번 구간 순수익: {self._usd_krw(self._hour_pnl)}  <i>(BNB 수수료 포함)</i>\n"
            f"⚡ 이번 구간 조기청산: {self._hour_abort_count}회\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 이번 구간 프리미엄\n"
            f"  최고: {max(self._hour_max_dc, self._hour_max_dt):.4%} / 기준: {self.cfg.dc_entry_threshold_rate:.4%}\n"
            f"  수수료 스택: {fee_stack_threshold:.4%}  (Leg B 조기청산 기준)\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📈 누적 거래: {self.trade_count}회\n"
            f"💰 누적 순수익: {self._usd_krw(self.total_pnl_usd)}  <i>(BNB 수수료 포함)</i>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"💼 잔고\n"
            f"  USDT: {self.usdt:,.2f}  USDC: {self.usdc:,.2f}\n"
            f"  BNB: {self.bnb:.6f}\n"
            f"  총자산: {self._eq_krw(equity)}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"⚠️ 이번구간: 조기청산 {self._period_abort_count}회  비상청산 {self._period_emergency_count}회  IOC miss {self._period_ioc_miss}회\n"
            f"   누  적: 조기청산 {self.abort_count}회  비상청산 {self.emergency_count}회  IOC miss {self.ioc_miss_count}회"
        )
        self._send_csv(
            f"📎 거래기록 ({span_min}분 주기, 누적 {self.trade_count}건)\n"
            f"누적 순수익: {self.total_pnl_usd:+.4f} USDT (BNB 수수료 포함)"
        )
        self.reset_period_stats()

    # ── IOC miss 디버깅 로그 ──────────────────────────────────────────────

    # miss_reason 분류:
    #   "EXPIRED"       - IOC 주문이 즉시 EXPIRED (ask 이미 상승, 버퍼 부족)
    #   "CANCELED"      - 거래소/WS 측 취소
    #   "TIMEOUT"       - leg_timeout_sec 초과 + WS 이벤트 미수신
    #   "SPEC_EXPIRED"  - Speculative 모드 Leg A EXPIRED
    #   "SPEC_CANCELED" - Speculative 모드 Leg A CANCELED
    #   "SPEC_TIMEOUT"  - Speculative 모드 Leg A 타임아웃

    def _log_ioc_miss(self, miss_reason: str, detect_method: str) -> None:
        """IOC miss 1건을 내부 리스트에 기록. 1시간마다 Excel로 플러시."""
        now_dt = datetime.now(tz=_KST)
        # 1시간 구간 라벨
        period_label = f"{now_dt.strftime('%Y-%m-%d')} {now_dt.hour:02d}:00~{now_dt.hour:02d}:59(KST)"

        elapsed_ms = (time.time() - self._leg_a_submit_ts) * 1000
        raw_ask    = self._leg_a_raw_ask
        ioc_price  = self.leg_a_price
        buf_rate   = self.cfg.ioc_price_buffer_rate
        buf_gap_pct = ((ioc_price - raw_ask) / raw_ask * 100) if raw_ask > 0 else 0.0

        row = {
            # Step 1: 시그널
            "S1_시각(KST)":         now_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "S1_1시간구간":         period_label,
            "S1_코인":              self.cfg.symbol,
            "S1_방향":              self.current_action.value,  # DT / DC
            "S1_시그널프리미엄(%)": round(self.entry_premium * 100, 5),
            "S1_진입임계값(%)":     round(self.cfg.dt_entry_threshold_rate * 100, 4),
            # Step 2: 주문 생성
            "S2_원본ask가격":       raw_ask,
            "S2_버퍼율(%)":         round(buf_rate * 100, 4),
            "S2_IOC주문가격":       ioc_price,
            "S2_버퍼갭(%)":         round(buf_gap_pct, 5),
            "S2_수량":              self.leg_a_btc_qty,
            "S2_notional(USD)":     round(self.leg_a_notional, 4),
            # Step 3: 주문 제출
            "S3_제출방식":          "WS" if self._leg_a_via_ws else "REST",
            "S3_주문ID":            self.leg_a_order_id,
            # Step 4: 체결 확인
            "S4_감지방식":          detect_method,  # "WS이벤트" / "REST폴백" / "타임아웃"
            "S4_경과시간(ms)":      round(elapsed_ms, 1),
            # Step 5: Miss 판정
            "S5_Miss사유":          miss_reason,
            "S5_누적miss수":        self.ioc_miss_count,
        }
        self._ioc_miss_rows.append(row)
        log.debug(f"[IOC_MISS_DEBUG] {miss_reason}  elapsed={elapsed_ms:.0f}ms  ask={raw_ask}  ioc={ioc_price}  buf={buf_rate:.4%}")

        # 메인 거래 CSV에도 IOC_MISS 행 기록 (순수익=0, 누계 불변)
        if self._csv_path:
            is_dt = self.current_action == ActionV2.TRADE_DT
            sym_a = self.cfg.binance_symbol_usdc if is_dt else self.cfg.binance_symbol_usdt
            equity = self.usdt + self.usdc
            miss_row = {
                "거래시각(KST)":   now_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "거래번호":        "",
                "방향":            "IOC_MISS",
                "신호프리미엄(%)": f"{self.signal_premium * 100:.4f}",
                "진입프리미엄(%)": f"{self.entry_premium * 100:.4f}",
                "매수시장":        sym_a,
                "매수가격(USD)":   f"{ioc_price:.6f}",
                "매도시장":        "",
                "매도가격(USD)":   "",
                "코인수량":        f"{self.leg_a_btc_qty:.8f}",
                "투입금액(USD)":   f"{self.leg_a_notional:.4f}",
                "수취금액(USD)":   "0.0000",
                "순수익(USD)":     "0.0000",
                "순수익(원)":      "0",
                "수수료추정(USD)": "0.0000",
                "체결소요시간(초)": f"{elapsed_ms / 1000:.1f}",
                "거래후USDT":      f"{self.usdt:.4f}",
                "거래후USDC":      f"{self.usdc:.4f}",
                "총자산(USD)":     f"{equity:.4f}",
                "총자산(원)":      f"{equity * self.fx_krw:.0f}",
                "누적순수익(USD)": f"{self.total_pnl_usd:.4f}",
                "누적순수익(원)":  f"{self.total_pnl_usd * self.fx_krw:.0f}",
                "잔여BNB":         f"{self.bnb:.6f}",
                "비고":            f"IOC_MISS:{miss_reason}",
            }
            try:
                with open(self._csv_path, "a", encoding="utf-8-sig", newline="") as f:
                    csv.DictWriter(f, fieldnames=_CSV_FIELDS).writerow(miss_row)
            except Exception as _e:
                log.warning(f"IOC_MISS CSV 기록 실패: {_e}")

    def _export_ioc_miss_excel(self) -> None:
        """누적된 IOC miss 행을 Excel에 기록 후 텔레그램 전송. 1시간마다 호출."""
        if not self._ioc_miss_rows:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            log.warning("openpyxl 없음 — IOC miss Excel 출력 스킵 (pip install openpyxl)")
            return

        wb = openpyxl.Workbook()
        ws_sheet = wb.active
        ws_sheet.title = "IOC_MISS"

        # ── 헤더 그룹 행 ──────────────────────────────────────────────────
        step_groups = [
            ("Step 1: 시그널", ["S1_시각(KST)", "S1_1시간구간", "S1_코인", "S1_방향",
                               "S1_시그널프리미엄(%)", "S1_진입임계값(%)"]),
            ("Step 2: 주문생성", ["S2_원본ask가격", "S2_버퍼율(%)", "S2_IOC주문가격",
                                  "S2_버퍼갭(%)", "S2_수량", "S2_notional(USD)"]),
            ("Step 3: 주문제출", ["S3_제출방식", "S3_주문ID"]),
            ("Step 4: 체결확인", ["S4_감지방식", "S4_경과시간(ms)"]),
            ("Step 5: Miss판정", ["S5_Miss사유", "S5_누적miss수"]),
        ]
        all_cols = [c for _, cols in step_groups for c in cols]

        group_colors = ["FFC7CE", "FFEB9C", "C6EFCE", "BDD7EE", "E2EFDA"]

        # 그룹 헤더 행 (1행)
        col_idx = 1
        for (label, cols), color in zip(step_groups, group_colors):
            start_col = col_idx
            end_col   = col_idx + len(cols) - 1
            ws_sheet.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            cell = ws_sheet.cell(row=1, column=start_col, value=label)
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor=color)
            col_idx += len(cols)

        # 컬럼명 행 (2행)
        for ci, col_name in enumerate(all_cols, start=1):
            cell = ws_sheet.cell(row=2, column=ci, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # 데이터 행 (3행~)
        miss_reason_colors = {
            "EXPIRED":       "FFC7CE",
            "CANCELED":      "FFEB9C",
            "TIMEOUT":       "D9D9D9",
            "SPEC_EXPIRED":  "FF9999",
            "SPEC_CANCELED": "FFD966",
            "SPEC_TIMEOUT":  "C0C0C0",
        }
        for ri, row_data in enumerate(self._ioc_miss_rows, start=3):
            reason = row_data.get("S5_Miss사유", "")
            row_color = miss_reason_colors.get(reason, "FFFFFF")
            for ci, col_name in enumerate(all_cols, start=1):
                cell = ws_sheet.cell(row=ri, column=ci, value=row_data.get(col_name, ""))
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.alignment = Alignment(horizontal="center")

        # 컬럼 너비 자동 조정
        for ci, col_name in enumerate(all_cols, start=1):
            ws_sheet.column_dimensions[get_column_letter(ci)].width = max(12, len(col_name) + 2)

        # 집계 시트 추가
        ws_agg = wb.create_sheet("20분집계")
        agg_header = ["1시간구간", "코인", "EXPIRED", "CANCELED", "TIMEOUT",
                      "SPEC_EXPIRED", "SPEC_CANCELED", "SPEC_TIMEOUT", "합계"]
        for ci, h in enumerate(agg_header, start=1):
            cell = ws_agg.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="BDD7EE")

        # 구간×코인별 집계
        from collections import defaultdict
        agg: dict = defaultdict(lambda: defaultdict(int))
        for row_data in self._ioc_miss_rows:
            key = (row_data["S1_1시간구간"], row_data["S1_코인"])
            agg[key][row_data["S5_Miss사유"]] += 1

        reasons = ["EXPIRED", "CANCELED", "TIMEOUT", "SPEC_EXPIRED", "SPEC_CANCELED", "SPEC_TIMEOUT"]
        for ri, ((period, coin), counts) in enumerate(sorted(agg.items()), start=2):
            total = sum(counts.values())
            ws_agg.cell(row=ri, column=1, value=period)
            ws_agg.cell(row=ri, column=2, value=coin)
            for ci, r in enumerate(reasons, start=3):
                ws_agg.cell(row=ri, column=ci, value=counts.get(r, 0))
            ws_agg.cell(row=ri, column=9, value=total)
            if total >= 5:
                for ci in range(1, 10):
                    ws_agg.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="FFC7CE")

        wb.save(self._ioc_miss_xlsx_path)
        log.info(f"IOC miss Excel 저장: {self._ioc_miss_xlsx_path}  ({len(self._ioc_miss_rows)}건)")
        # 텔레그램 파일 전송
        try:
            caption = (
                f"📊 IOC miss 디버그 [{self.cfg.symbol}] "
                f"— 전체 {len(self._ioc_miss_rows)}건 / 구간 {self._period_ioc_miss}건"
            )
            send_telegram_document(self._ioc_miss_xlsx_path, caption=caption)
        except Exception as e:
            log.warning(f"IOC miss Excel 텔레그램 전송 실패: {e}")

    def _check_ioc_summary(self) -> None:
        """1시간마다 IOC miss 현황 텔레그램 전송 + Excel 생성."""
        now = time.time()
        if now - self._last_ioc_summary_ts < self._ioc_summary_sec:
            return
        if not self.suppress_hourly_telegram:
            spec_mode = "SPEC" if self.cfg.use_speculative_leg_b else "일반"
            span_min_ioc = max(1, int(round(self._ioc_summary_sec / 60.0)))
            self._notify(
                f"🎯 <b>IOC miss 현황 [{self.cfg.symbol}] ({span_min_ioc}분단위)</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"구간 miss: {self._period_ioc_miss}회  |  누적 miss: {self.ioc_miss_count}회\n"
                f"구간 거래: {self._hour_trade_count}회\n"
                f"miss율: {self._period_ioc_miss / max(1, self._period_ioc_miss + self._hour_trade_count):.1%}\n"
                f"모드: {spec_mode}  임계값: {self.cfg.dt_entry_threshold_rate:.4%}"
            )
        if not self.suppress_hourly_telegram:
            self._export_ioc_miss_excel()
            self._export_event_excel()
        self._period_ioc_miss = 0
        self._last_ioc_summary_ts = now

    # ── 통합 이벤트 로그 ──────────────────────────────────────────────────

    def _log_event(self, ev_type: str, direction: str, entry_premium: float,
                   current_premium: float, threshold: float,
                   elapsed_ms: float, note: str = "") -> None:
        """모든 이벤트(IOC miss / 조기청산 / 타임아웃)를 통합 이벤트 로그에 기록."""
        now_dt = datetime.now(tz=_KST)
        period_label = f"{now_dt.strftime('%Y-%m-%d')} {now_dt.hour:02d}:00~{now_dt.hour:02d}:59(KST)"
        self._event_rows.append({
            "시각(KST)":        now_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "20분구간":         f"{now_dt.strftime('%Y-%m-%d')} {now_dt.hour:02d}:{(now_dt.minute//20)*20:02d}",
            "코인":             self.cfg.symbol,
            "이벤트유형":       ev_type,
            "방향":             direction,
            "진입프리미엄(%)":  round(entry_premium * 100, 4),
            "현재프리미엄(%)":  round(current_premium * 100, 4),
            "임계값(%)":        round(threshold * 100, 4),
            "경과시간(ms)":     round(elapsed_ms, 1),
            "비고":             note,
        })

    def _export_event_excel(self) -> None:
        """누적 이벤트 행을 Excel로 저장 후 텔레그램 전송."""
        if not self._event_rows:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            log.warning("openpyxl 없음 — 이벤트 Excel 출력 스킵")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "이벤트로그"

        cols = ["시각(KST)", "20분구간", "코인", "이벤트유형", "방향",
                "진입프리미엄(%)", "현재프리미엄(%)", "임계값(%)", "경과시간(ms)", "비고"]
        ev_colors = {
            "IOC_MISS":     "FFC7CE",
            "SPEC_MISS":    "FF9999",
            "TIMEOUT_A":    "D9D9D9",
            "TIMEOUT_B":    "C0C0C0",
            "SPEC_TIMEOUT": "AAAAAA",
            "ABORT":        "FFEB9C",
        }
        # 헤더
        for ci, h in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")

        for ri, row in enumerate(self._event_rows, start=2):
            color = ev_colors.get(row.get("이벤트유형", ""), "FFFFFF")
            for ci, col in enumerate(cols, start=1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
                cell.fill = PatternFill("solid", fgColor=color)

        # 열 너비 자동
        for ci, col in enumerate(cols, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(col) + 2)

        # 20분 집계 시트
        ws2 = wb.create_sheet("20분집계")
        agg_cols = ["20분구간", "코인", "IOC_MISS", "SPEC_MISS", "TIMEOUT_A", "TIMEOUT_B", "SPEC_TIMEOUT", "ABORT", "합계"]
        for ci, h in enumerate(agg_cols, start=1):
            ws2.cell(row=1, column=ci, value=h).font = Font(bold=True)
        from collections import defaultdict
        agg: dict = defaultdict(lambda: defaultdict(int))
        for row in self._event_rows:
            key = (row["20분구간"], row["코인"])
            agg[key][row["이벤트유형"]] += 1
        ev_types = ["IOC_MISS", "SPEC_MISS", "TIMEOUT_A", "TIMEOUT_B", "SPEC_TIMEOUT", "ABORT"]
        for ri, ((period, coin), counts) in enumerate(sorted(agg.items()), start=2):
            total = sum(counts.values())
            ws2.cell(row=ri, column=1, value=period)
            ws2.cell(row=ri, column=2, value=coin)
            for ci, t in enumerate(ev_types, start=3):
                ws2.cell(row=ri, column=ci, value=counts.get(t, 0))
            ws2.cell(row=ri, column=9, value=total)

        wb.save(self._event_xlsx_path)
        log.info(f"이벤트 Excel 저장: {self._event_xlsx_path} ({len(self._event_rows)}건)")
        try:
            caption = (
                f"📋 이벤트 디버그 [{self.cfg.symbol}] "
                f"— 전체 {len(self._event_rows)}건"
            )
            send_telegram_document(self._event_xlsx_path, caption=caption)
        except Exception as e:
            log.warning(f"이벤트 Excel 텔레그램 전송 실패: {e}")

    # ── 정기 잔고 동기화 ──────────────────────────────────────────────────

    _PERIODIC_SYNC_INTERVAL: float = 3600.0  # 1시간

    def _check_periodic_sync(self) -> None:
        """1시간마다 거래소 잔고를 강제 동기화 (거래 없어도 실행)."""
        if self.dry_run:
            return
        now = time.time()
        if now - self._last_periodic_sync_ts >= self._PERIODIC_SYNC_INTERVAL:
            log.info(f"[{self.cfg.symbol}] 정기 잔고 동기화")
            self.sync_balances()
            self._last_periodic_sync_ts = now

    # ── 메인 루프 ─────────────────────────────────────────────────────────

    def run(self) -> None:
        # 시작 알림은 HTML 미사용 — %·기호가 parse 오류 나면 본문이 통째로 안 보일 수 있음.
        mode = "DRY RUN" if self.dry_run else "LIVE"
        fee_stack = self.cfg.fee_rate + self._eff_fee() + 2.0 * self.cfg.slippage_rate
        sym = self.cfg.symbol
        ws_price_mode = "WS event-driven" if self.price_feed      else "REST polling"
        ws_order_mode = "WS Order API"    if self.ws_order_client else "REST"
        spec_mode_str = "Speculative (Leg A+B 동시 제출)" if self.cfg.use_speculative_leg_b else "Sequential (Leg A → Leg B)"
        abort_thr = self.cfg.leg_b_abort_threshold_rate
        abort_str = f"{abort_thr:.4%}" if abort_thr is not None else f"{fee_stack:.4%}(fee_stack)"
        if not self.suppress_hourly_telegram:
            self._notify(
                f"🚀 v1.28 라이브 엔진 시작 ({mode}) [{sym}]\n\n"
                f"[v1.28 업데이트]\n"
                f"  · 통합 풀 모드: TRX + DOGE + XRP 단일 자본 풀 운영\n"
                f"  · GlobalLock: 1개 코인 거래 중 나머지 진입 차단 (선착순)\n"
                f"  · 풀 자본: 644 USDT + 644 USDC / 거래당 notional 322 USDT\n\n"
                f"코인: {sym}  ({self.cfg.binance_symbol_usdt} / {self.cfg.binance_symbol_usdc})\n"
                f"풀 자본: USDT={self.cfg.initial_usdt:.0f}  USDC={self.cfg.initial_usdc:.0f}\n"
                f"{'🔒 GlobalLock ON' if self.global_lock is not None else '🔓 GlobalLock OFF'}\n\n"
                f"진입 임계값: DT {self.cfg.dt_entry_threshold_rate:.4%}  DC {self.cfg.dc_entry_threshold_rate:.4%}\n"
                f"실행 모드: {spec_mode_str}\n"
                f"Abort 임계값: {abort_str}\n"
                f"수수료 스택: {fee_stack:.4%}  (레그당 {self._eff_fee():.4%} × 2)\n\n"
                f"가격·체결 감지: {ws_price_mode}  |  주문: {ws_order_mode}\n"
                f"timeout={self.leg_timeout_sec}s  CSV: {self._csv_path}",
                telegram_parse_mode=None,
            )
        # IOC/이벤트 요약 주기를 운영요약 주기와 동기화
        self._ioc_summary_sec = self.telegram_summary_sec
        # 8코인 동시 정기동기화 방지: 심볼별 오프셋으로 분산 (7분 간격)
        _SYNC_SYMBOLS = ["TRX", "DOGE", "XRP", "SOL", "BNB", "ADA", "LINK", "AVAX"]
        _sync_offset = _SYNC_SYMBOLS.index(self.cfg.symbol) * 420 if self.cfg.symbol in _SYNC_SYMBOLS else 0
        self._last_periodic_sync_ts = time.time() - self._PERIODIC_SYNC_INTERVAL + _sync_offset
        self.sync_balances()

        # ── 시작 시 고아 코인 감지 및 즉시 청산 (BUG-24) ──────────────────
        # 재시작 중 Leg A 체결 후 Leg B 미처리로 코인이 free 상태로 남을 수 있음.
        # 시작 직후 해당 코인 잔고를 확인하고 min_qty 이상이면 USDC로 즉시 시장가 매도.
        if not self.dry_run and self.cfg.symbol != "BNB":
            # BNB 제외: 수수료 코인이라 min_reserve 초과분이 항상 고아가 아님.
            # BNB 거래 고아는 별도 오픈오더 확인이 필요 — 단순 잔고 차이로 판단 불가.
            _start_balances = self.binance.get_spot_balances()
            _coin_qty = _start_balances.get(self.cfg.symbol, 0.0)
            _sell_qty = self.binance.normalize_market_quantity(
                self.cfg.binance_symbol_usdc, _coin_qty
            )
            if _sell_qty > 0:
                log.warning(
                    f"[{self.cfg.symbol}] 시작 시 고아 코인 감지: "
                    f"{_coin_qty} {self.cfg.symbol} → USDC 즉시 시장가 매도"
                )
                self._notify(
                    f"⚠️ [{self.cfg.symbol}] 시작 시 고아 코인 청산\n"
                    f"{_coin_qty} {self.cfg.symbol} → USDC 시장가 매도"
                )
                try:
                    _result = self.binance.place_market_order(
                        self.cfg.binance_symbol_usdc, "SELL", _sell_qty
                    )
                    _received = float(_result.get("cummulativeQuoteQty", 0.0))
                    log.info(
                        f"[{self.cfg.symbol}] 고아 코인 청산 완료: "
                        f"{_received:.4f} USDC 수취"
                    )
                    self._notify(
                        f"✅ [{self.cfg.symbol}] 고아 코인 청산 완료: "
                        f"{_received:.4f} USDC 수취"
                    )
                    self.sync_balances()
                except Exception as _e:
                    log.error(f"[{self.cfg.symbol}] 고아 코인 청산 실패: {_e}")
                    self._notify(
                        f"❌ [{self.cfg.symbol}] 고아 코인 청산 실패: {_e}"
                    )

        # 재시작 후 잔고 쏠림 시 쿨다운 무시: 복원된 _last_rebalance_ts 때문에
        # 이전 세션 쿨다운이 유지되는 동안 잔고가 극단적으로 쏠려도 리밸런싱이 막히는 문제 방지.
        # (GlobalLock 공유 쿨다운은 인메모리라 재시작 시 이미 리셋됨 — 개별 타임스탬프만 문제)
        _total_on_start = self.usdt + self.usdc
        if _total_on_start > 1.0:
            _skew_on_start = max(self.usdt, self.usdc) / _total_on_start
            if _skew_on_start > self.cfg.rebalance_skew_threshold:
                log.info(
                    f"[{self.cfg.symbol}] 시작 잔고 쏠림({_skew_on_start:.1%}) 감지 "
                    f"→ _last_rebalance_ts 리셋 (즉시 리밸런싱 허용)"
                )
                self._last_rebalance_ts = 0.0

        # WS 콜백 등록 (event-driven 모드)
        event_driven = self.price_feed is not None
        if event_driven:
            self.price_feed.register_callbacks(
                on_price=self._on_price_update,
                on_fill=self._on_fill,
            )
            log.info(f"[{self.cfg.symbol}] event-driven 모드 (WS 콜백 등록)")
        else:
            log.info(f"[{self.cfg.symbol}] polling 모드 ({self.poll_interval_sec}s)")

        while True:
            try:
                self._check_hourly()
                self._check_ioc_summary()
                self._check_periodic_sync()
                if event_driven:
                    self._wake_event.wait(timeout=1.0)
                    self._wake_event.clear()
                else:
                    time.sleep(self.poll_interval_sec)
                self._tick()
            except KeyboardInterrupt:
                self._notify("⛔ 엔진 정지 (KeyboardInterrupt)")
                break
            except Exception as e:
                log.error(f"틱 오류: {e}", exc_info=True)
                if not event_driven:
                    time.sleep(self.poll_interval_sec)

    def _tick(self) -> None:
        if self.state == TradeState.IDLE:
            self._idle_tick()
        elif self.state == TradeState.LEG_A_PENDING:
            self._leg_a_tick()
        elif self.state == TradeState.LEG_B_PENDING:
            self._leg_b_tick()
        elif self.state == TradeState.SPEC_PENDING:
            self._spec_tick()

    # ── 상태 영속화 ───────────────────────────────────────────────────────

    def _load_state(self) -> None:
        """재시작 시 리밸런싱 쿨다운 상태 복원."""
        import json as _json
        p = Path(self._state_path)
        if not p.exists():
            return
        try:
            state = _json.loads(p.read_text(encoding="utf-8"))
            self._last_rebalance_ts = float(state.get("last_rebalance_ts", 0.0))
            log.info(
                f"상태 복원: last_rebalance_ts={self._last_rebalance_ts:.0f} "
                f"({self._now_str()})"
            )
        except Exception as e:
            log.warning(f"상태 파일 로드 실패 (무시): {e}")

    def _save_state(self) -> None:
        """리밸런싱 타임스탬프를 파일에 저장."""
        import json as _json
        try:
            Path(self._state_path).write_text(
                _json.dumps({"last_rebalance_ts": self._last_rebalance_ts}),
                encoding="utf-8",
            )
        except Exception as e:
            log.warning(f"상태 파일 저장 실패: {e}")

    # ── 리밸런싱 ──────────────────────────────────────────────────────────

    def _check_rebalance(self) -> None:
        """USDT/USDC 비율이 skew_threshold 초과 시 Convert로 50:50 복귀."""
        total = self.usdt + self.usdc
        if total < 1.0:
            return

        now = time.time()
        # GlobalLock 공유 쿨다운 체크: 다른 엔진이 이미 리밸런싱했으면 전체 스킵
        if self.global_lock is not None and self.global_lock.is_rebalance_cooldown():
            return
        if now - self._last_rebalance_ts < self.cfg.rebalance_min_interval_sec:
            return

        usdt_ratio = self.usdt / total
        usdc_ratio = self.usdc / total
        threshold = self.cfg.rebalance_skew_threshold

        if usdt_ratio <= threshold and usdc_ratio <= threshold:
            return  # 쏠림 없음 (내부 잔고 기준)

        # 쏠림 감지 시 실잔고로 재확인: 다른 엔진의 리밸런싱 성공 후 내부 잔고가 스테일일 수 있음
        self.sync_balances()
        total = self.usdt + self.usdc
        if total < 1.0:
            return
        usdt_ratio = self.usdt / total
        usdc_ratio = self.usdc / total
        if usdt_ratio <= threshold and usdc_ratio <= threshold:
            log.debug(f"[{self.cfg.symbol}] 리밸런싱 스킵 — 실잔고 재확인 후 쏠림 없음")
            return  # 실잔고는 균형 상태

        # 50:50 목표: 더 많은 쪽에서 절반만큼을 반대쪽으로 Convert
        if usdt_ratio > threshold:
            from_asset, to_asset = "USDT", "USDC"
            from_amount = round(self.usdt - total * 0.5, 4)
        else:
            from_asset, to_asset = "USDC", "USDT"
            from_amount = round(self.usdc - total * 0.5, 4)

        if from_amount < 1.0:
            return

        # GlobalTradeLock: 다른 코인이 거래 또는 리밸런싱 중이면 이번 틱 스킵
        # → 4코인이 동시에 Convert를 발사하는 문제 방지
        lock_key = f"{self.cfg.symbol}:REB"
        if self.global_lock is not None:
            if not self.global_lock.try_acquire(lock_key):
                log.debug(
                    f"[{self.cfg.symbol}] 리밸런싱 스킵 — GlobalLock 홀딩 중 "
                    f"(holder={self.global_lock.holder})"
                )
                return

        try:
            log.info(
                f"[{self._now_str()}] 리밸런싱: {from_asset}→{to_asset}  "
                f"amount={from_amount:.2f}  USDT={self.usdt:.2f}  USDC={self.usdc:.2f}"
            )
            _last_sym = self.global_lock.last_trade_symbol if self.global_lock else ""
            _last_dir = self.global_lock.last_trade_dir if self.global_lock else ""
            _reb_no = (self.global_lock.rebalance_count + 1) if self.global_lock else 0
            _trigger_info = f"{_last_sym} {_last_dir}" if _last_sym else "—"
            self._notify(
                f"⚖️ <b>v1.28 리밸런싱 #{_reb_no}</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"Spot {from_amount:.2f} {from_asset} → {to_asset}\n"
                f"현재: USDT={self.usdt:.2f}  USDC={self.usdc:.2f}\n"
                f"비율: USDT {self.usdt/total:.1%}  USDC {self.usdc/total:.1%}\n"
                f"트리거: {_trigger_info}"
            )

            if not self.dry_run:
                rebalance_ok = False
                method_used = "spot"
                # Convert는 ~0.125% 스프레드로 USDT+USDC 총액을 직접 감소시킴.
                # USDCUSDT 현물은 스프레드 ~0.001% + BNB 수수료만 발생 → 총액 변동 거의 없음.
                # → 현물을 기본으로 사용하고 Convert는 폴백으로.
                try:
                    if from_asset == "USDT":
                        # USDT → USDC: quoteOrderQty로 USDC 매수
                        result = self.binance.place_market_buy_quote(
                            "USDCUSDT", from_amount
                        )
                    else:
                        # USDC → USDT: 실 free 잔고 확인 후 USDC 매도
                        # 거래 직후 정산 딜레이로 self.usdc가 실제 free 잔고보다 클 수 있음
                        actual_free_usdc = self.binance.get_spot_balances().get("USDC", 0.0)
                        safe_amount = min(from_amount, actual_free_usdc)
                        log.info(
                            f"[{self._now_str()}] 리밸런싱 SELL: want={from_amount:.2f} USDC, "
                            f"free={actual_free_usdc:.2f} USDC, sell={safe_amount:.2f}"
                        )
                        qty = self.binance.normalize_market_quantity(
                            "USDCUSDT", safe_amount
                        )
                        if qty <= 0:
                            raise RuntimeError(
                                f"USDCUSDT SELL qty=0 "
                                f"(want={from_amount:.2f}, free_usdc={actual_free_usdc:.2f})"
                            )
                        result = self.binance.place_market_order(
                            "USDCUSDT", "SELL", qty
                        )
                    log.info(f"현물 리밸런싱 완료: {result}")
                    rebalance_ok = True
                except Exception as e_spot:
                    # 현물 실패 → 1분 쿨다운 후 재시도 (Convert API quoteId 버그로 폴백 비활성화)
                    # -2010 원인: 거래 직후 바이낸스 정산 딜레이로 free USDC 일시 부족
                    # 보통 1~2분 내 자연 해소됨
                    log.warning(
                        f"[{self._now_str()}] 현물 리밸런싱 실패: {e_spot} "
                        f"→ 1분 후 재시도"
                    )
                    self._notify(f"⚠️ 현물 리밸런싱 실패 (1분 후 재시도): {e_spot}")
                    self._last_rebalance_ts = now - self.cfg.rebalance_min_interval_sec + 60
                    if self.global_lock is not None:
                        self.global_lock.set_rebalance_cooldown(now + 60)

                if rebalance_ok:
                    # Spot 현물은 즉시 체결 → 즉시 sync로 실잔고 반영.
                    # (BUG-14는 Convert 딜레이 문제였음. Spot은 sync 안전)
                    # 다른 엔진의 스테일 내부 잔고도 sync 로 교정됨.
                    self.sync_balances()
                    self._last_rebalance_ts = now
                    # 성공 후 GlobalLock 공유 쿨다운: 다른 엔진이 스테일 상태로
                    # 즉시 이중 리밸런싱하는 것을 방지.
                    # rebalance_min_interval_sec=0이어도 최소 60s 쿨다운 보장.
                    # (60s 안에 모든 엔진이 _check_rebalance에서 실제 잔고 확인 가능)
                    _post_reb_cooldown = max(self.cfg.rebalance_min_interval_sec, 60.0)
                    _reb_count = 0
                    if self.global_lock is not None:
                        self.global_lock.set_rebalance_cooldown(
                            now + _post_reb_cooldown
                        )
                        _reb_count = self.global_lock.increment_rebalance_count()
                    self._save_state()
                    self._notify(
                        f"✅ <b>리밸런싱 완료 #{_reb_count}</b> ({method_used})\n"
                        f"USDT={self.usdt:.2f}  USDC={self.usdc:.2f}\n"
                        f"총자산={self._eq_krw(self.usdt + self.usdc)}"
                    )
            else:
                # dry_run: 잔고 직접 조정
                if from_asset == "USDT":
                    self.usdt -= from_amount
                    self.usdc += from_amount
                else:
                    self.usdc -= from_amount
                    self.usdt += from_amount
                self._last_rebalance_ts = now
                self._save_state()
                log.info(f"[DRY] 리밸런싱 후: USDT={self.usdt:.2f}  USDC={self.usdc:.2f}")
        finally:
            if self.global_lock is not None:
                self.global_lock.release(lock_key)

    # ── IDLE ──────────────────────────────────────────────────────────────

    def _idle_tick(self) -> None:
        self._check_rebalance()
        if self.cfg.min_bnb_fee_reserve > 0 and self.bnb < self.cfg.min_bnb_fee_reserve:
            log.debug(
                f"[IDLE] BNB {self.bnb:.4f} < 최소보유 {self.cfg.min_bnb_fee_reserve} → 진입 스킵"
            )
            return
        mid_ut, mid_uc = self._get_mids()
        # bid/ask 기반 실행가능 스프레드 계산 (진입 결정 및 시간대별 통계용)
        sym_ut = self.cfg.binance_symbol_usdt
        sym_uc = self.cfg.binance_symbol_usdc
        ba_ut = self._get_bid_ask(sym_ut)   # (bid_USDT, ask_USDT)
        ba_uc = self._get_bid_ask(sym_uc)   # (bid_USDC, ask_USDC)
        exec_dc = (ba_uc[0] - ba_ut[1]) / ba_ut[1] if ba_ut[1] > 0 else 0.0  # DC: bid_USDC - ask_USDT
        exec_dt = (ba_ut[0] - ba_uc[1]) / ba_uc[1] if ba_uc[1] > 0 else 0.0  # DT: bid_USDT - ask_USDC
        # 방향별 최고 프리미엄 독립 추적
        self._hour_max_dc = max(self._hour_max_dc, exec_dc)
        self._hour_max_dt = max(self._hour_max_dt, exec_dt)
        log.debug(
            f"[IDLE] exec_dc={exec_dc:.5%}  exec_dt={exec_dt:.5%}  "
            f"USDT={self.usdt:.2f}  USDC={self.usdc:.2f}"
        )
        decision = decide_v2(
            mid_usdt=mid_ut, mid_usdc=mid_uc,
            usdt=self.usdt, usdc=self.usdc,
            dt_entry_threshold_rate=self.cfg.dt_entry_threshold_rate,
            dc_entry_threshold_rate=self.cfg.dc_entry_threshold_rate,
            fee_rate=self._eff_fee(),
            slippage_rate=self.cfg.slippage_rate,
            entry_split_fraction=self.cfg.entry_split_fraction,
            exec_dc_premium=exec_dc,
            exec_dt_premium=exec_dt,
        )
        if decision.action in (ActionV2.TRADE_DT, ActionV2.TRADE_DC):
            # 통합 풀 모드: 다른 코인이 거래 중이면 이번 틱 스킵
            if self.global_lock is not None and not self.global_lock.try_acquire(self.cfg.symbol):
                return

            if decision.action == ActionV2.TRADE_DT:
                if self.cfg.use_speculative_leg_b:
                    self._start_speculative(mid_ut, mid_uc, ActionV2.TRADE_DT)
                else:
                    self._start_leg_a(mid_ut, mid_uc, ActionV2.TRADE_DT)
            else:
                if self.cfg.use_speculative_leg_b:
                    self._start_speculative(mid_ut, mid_uc, ActionV2.TRADE_DC)
                else:
                    self._start_leg_a(mid_ut, mid_uc, ActionV2.TRADE_DC)

            # 잔고 부족 등으로 진입 실패 시(state 여전히 IDLE) 즉시 락 해제
            if self.global_lock is not None and self.state == TradeState.IDLE:
                self.global_lock.release(self.cfg.symbol)

    # ── Leg A 진입 ────────────────────────────────────────────────────────

    def _start_leg_a(self, mid_ut: float, mid_uc: float, action: ActionV2) -> None:
        is_dt = action == ActionV2.TRADE_DT
        symbol = self.cfg.binance_symbol_usdc if is_dt else self.cfg.binance_symbol_usdt
        stable_name = "USDC" if is_dt else "USDT"

        required_stable = self.usdc if is_dt else self.usdt
        notional = (self.usdt + self.usdc) * self.cfg.entry_split_fraction
        # available 잔고를 초과하지 않도록 notional 상한 적용
        # 기존 required_stable < notional * 0.99 체크는 1% 오차를 허용했으나
        # 실제 주문 시 -2010(잔고 부족) 반복 유발 → notional을 잔고로 clamp
        if required_stable < 10.0:
            log.debug(f"{'DT' if is_dt else 'DC'} 스킵: {stable_name} 잔고 없음 "
                      f"({required_stable:.2f})")
            return
        notional = min(notional, required_stable)

        raw_ask = self._get_best_ask(symbol)
        # IOC 버퍼: WS→거래소 레이턴시(5~15ms) 동안 ask 이동 커버
        buffered_ask = raw_ask * (1.0 + self.cfg.ioc_price_buffer_rate) if self.cfg.ioc_price_buffer_rate > 0 else raw_ask
        buy_price = self.binance.normalize_price_ceil(symbol, buffered_ask)
        btc_qty = notional / (buy_price * (1.0 + self._eff_fee()))
        btc_qty = self.binance.normalize_market_quantity(symbol, btc_qty)
        if btc_qty <= 0:
            return

        # 진입프리미엄: bid/ask 기반 실행가능 스프레드 (LegA ask vs LegB bid)
        sym_legb = self.cfg.binance_symbol_usdt if is_dt else self.cfg.binance_symbol_usdc
        bid_legb = self._get_best_bid(sym_legb)
        premium = (bid_legb - raw_ask) / raw_ask if raw_ask > 0 else 0.0

        # 가격 재검증: decide_v2 이후 WS 가격 갱신으로 premium이 임계값 아래로 내려갔으면 스킵
        threshold = self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate
        if premium < threshold:
            log.info(
                f"[{self._now_str()}] {'DT' if is_dt else 'DC'} Leg A 스킵 (가격 역전): "
                f"signal_premium={premium:.4%} < threshold={threshold:.4%}"
            )
            return

        # 수익성 체크: IOC는 buy_price(상한)가 아닌 raw_ask에 체결되므로 raw_ask 기준으로 계산
        # Leg A는 IOC(taker), Leg B는 GTC(maker) → fee_stack = fee_rate + fee_maker_rate
        actual_fee_stack = self.cfg.fee_rate + self._eff_fee() + 2.0 * self.cfg.slippage_rate
        actual_premium = (bid_legb - raw_ask) / raw_ask if raw_ask > 0 else 0.0
        if actual_premium < actual_fee_stack:
            log.info(
                f"[{self._now_str()}] {'DT' if is_dt else 'DC'} Leg A 스킵 (수익성 없음): "
                f"actual_premium={actual_premium:.4%} < fee_stack={actual_fee_stack:.4%}  "
                f"(signal={premium:.4%}  buf={self.cfg.ioc_price_buffer_rate:.4%})"
            )
            return

        log.info(
            f"[{self._now_str()}] {'DT' if is_dt else 'DC'} Leg A: "
            f"BUY {btc_qty:.8f} {self.cfg.symbol} @ {buy_price:.2f} {stable_name}  "
            f"(ask={raw_ask:.8f} buf={self.cfg.ioc_price_buffer_rate:.4%})  notional={notional:.2f}  exec_premium={premium:.4%}"
        )

        self._last_leg_a_attempt_ts = time.time()
        # 디버깅용 상태 저장
        self._leg_a_raw_ask = raw_ask
        self._leg_a_submit_ts = time.time()
        if self.dry_run:
            self.leg_a_order_id = "DRY"
            self._leg_a_via_ws = False
        else:
            try:
                via_ws = bool(self.ws_order_client and self.ws_order_client.is_ready())
                self._leg_a_via_ws = via_ws
                if via_ws:
                    resp = self.ws_order_client.place_limit_order(
                        symbol, "BUY", btc_qty, buy_price, time_in_force="IOC"
                    )
                else:
                    resp = self.binance.place_limit_order(
                        symbol, "BUY", btc_qty, buy_price, time_in_force="IOC"
                    )
                self.leg_a_order_id = str(resp["orderId"])
            except Exception as e:
                log.error(f"Leg A 주문 실패: {e}")
                # [BUG-6] -2010 잔고 부족: 다른 엔진이 거래해 실잔고 감소 → 즉시 재동기화
                # 다음 틱부터 정확한 잔고로 진입 판단 가능
                err_str = str(e)
                if "-2010" in err_str or "insufficient balance" in err_str.lower():
                    try:
                        self.sync_balances()
                        log.info(f"[BUG-6] -2010 감지 → 잔고 재동기화 완료")
                    except Exception as e_sync:
                        log.warning(f"잔고 재동기화 실패: {e_sync}")
                return

        self.current_action = action
        self.leg_a_notional = notional
        self.leg_a_btc_qty = btc_qty
        self.leg_a_price = buy_price
        self.signal_premium = premium   # 신호 시점 raw_ask 기준 (threshold 판단값)
        self.entry_premium = premium    # 체결 후 실체결가 기준으로 덮어씌워짐
        self.state = TradeState.LEG_A_PENDING
        self.order_time = time.time()

    # ── Speculative 진입 (Leg A + Leg B 동시 제출) ───────────────────────

    def _start_speculative(self, mid_ut: float, mid_uc: float, action: ActionV2) -> None:
        """Leg A IOC + Leg B GTC 동시 제출. Leg A miss 시 Leg B 즉시 취소."""
        is_dt = action == ActionV2.TRADE_DT
        sym_a = self.cfg.binance_symbol_usdc if is_dt else self.cfg.binance_symbol_usdt
        sym_b = self.cfg.binance_symbol_usdt if is_dt else self.cfg.binance_symbol_usdc
        required_stable = self.usdc if is_dt else self.usdt
        stable_name = "USDC" if is_dt else "USDT"

        notional = (self.usdt + self.usdc) * self.cfg.entry_split_fraction
        if required_stable < notional * 0.99:
            log.debug(f"{'DT' if is_dt else 'DC'} Spec 스킵: {stable_name} 부족")
            return

        # Leg A: IOC 매수 가격
        raw_ask = self._get_best_ask(sym_a)
        buffered_ask = raw_ask * (1.0 + self.cfg.ioc_price_buffer_rate) if self.cfg.ioc_price_buffer_rate > 0 else raw_ask
        buy_price = self.binance.normalize_price_ceil(sym_a, buffered_ask)
        btc_qty = notional / (buy_price * (1.0 + self._eff_fee()))
        btc_qty = self.binance.normalize_market_quantity(sym_a, btc_qty)
        if btc_qty <= 0:
            return

        # Leg B: GTC 매도 가격 (현재 bid)
        sell_price = self.binance.normalize_price(sym_b, self._get_best_bid(sym_b))

        # 진입프리미엄: bid/ask 기반 실행가능 스프레드
        premium = (sell_price - buy_price) / buy_price if buy_price > 0 else 0.0
        log.info(
            f"[{self._now_str()}] {'DT' if is_dt else 'DC'} SPEC: "
            f"BUY {btc_qty:.8f} {self.cfg.symbol} @ {buy_price:.2f} {stable_name} (IOC)  "
            f"SELL @ {sell_price:.2f} (GTC)  exec_premium={premium:.4%}"
        )

        self._last_leg_a_attempt_ts = time.time()
        # 디버깅용 상태 저장
        self._leg_a_raw_ask = raw_ask
        self._leg_a_submit_ts = time.time()

        if self.dry_run:
            self.leg_a_order_id = "DRY_A"
            self.leg_b_order_id = "DRY_B"
            self._leg_a_via_ws = False
        else:
            # Leg A 제출
            try:
                via_ws = bool(self.ws_order_client and self.ws_order_client.is_ready())
                self._leg_a_via_ws = via_ws
                if via_ws:
                    resp_a = self.ws_order_client.place_limit_order(sym_a, "BUY", btc_qty, buy_price, time_in_force="IOC")
                else:
                    resp_a = self.binance.place_limit_order(sym_a, "BUY", btc_qty, buy_price, time_in_force="IOC")
                self.leg_a_order_id = str(resp_a["orderId"])
            except Exception as e:
                log.error(f"SPEC Leg A 주문 실패: {e}")
                return

            # Leg B 즉시 제출 (Leg A 결과 기다리지 않음)
            try:
                if self.ws_order_client and self.ws_order_client.is_ready():
                    resp_b = self.ws_order_client.place_limit_order(sym_b, "SELL", btc_qty, sell_price, time_in_force="GTC")
                else:
                    resp_b = self.binance.place_limit_order(sym_b, "SELL", btc_qty, sell_price, time_in_force="GTC")
                self.leg_b_order_id = str(resp_b["orderId"])
            except Exception as e:
                log.error(f"SPEC Leg B 주문 실패: {e}")
                # Leg B 실패 → Leg A만 있으므로 일반 LEG_A_PENDING으로 계속
                self.leg_a_order_id = self.leg_a_order_id
                self.current_action = action
                self.leg_a_notional = notional
                self.leg_a_btc_qty = btc_qty
                self.leg_a_price = buy_price
                self.leg_b_price = sell_price
                self.entry_premium = premium
                self.state = TradeState.LEG_A_PENDING
                self.order_time = time.time()
                return

        self.current_action = action
        self.leg_a_notional = notional
        self.leg_a_btc_qty = btc_qty
        self.leg_a_price = buy_price
        self.leg_b_price = sell_price
        self.entry_premium = premium
        self._spec_leg_a_filled = False
        self._spec_leg_b_filled = False
        self._spec_received = 0.0
        self.state = TradeState.SPEC_PENDING
        self.order_time = time.time()

    # ── Leg A 모니터링 ────────────────────────────────────────────────────

    def _leg_a_tick(self) -> None:
        elapsed = time.time() - self.order_time
        is_dt = self.current_action == ActionV2.TRADE_DT
        symbol = self.cfg.binance_symbol_usdc if is_dt else self.cfg.binance_symbol_usdt

        if elapsed >= self.leg_timeout_sec:
            log.info(f"[{self._now_str()}] Leg A 타임아웃 ({elapsed:.1f}s) — IOC 자동취소됨, IDLE 복귀")
            if not self.dry_run:
                # IOC는 이미 자동취소됨. WS 이벤트 누락 시 REST로 상태 재확인만
                try:
                    status = self.binance.get_order_status(symbol, self.leg_a_order_id)
                    if status.get("status") == "FILLED":
                        exec_qty = float(status.get("executedQty", 0.0))
                        self.leg_a_btc_qty   = exec_qty
                        self.leg_a_fill_time = time.time()
                        self._last_rest_check_ts = 0.0
                        self.leg_a_notional = float(status.get("cummulativeQuoteQty", 0.0)) or self.leg_a_notional
                        if is_dt:
                            self.usdc = max(self.usdc - self.leg_a_notional, 0.0)
                        else:
                            self.usdt = max(self.usdt - self.leg_a_notional, 0.0)
                        self._place_leg_b()
                        return
                except Exception:
                    pass
            self.ioc_miss_count += 1
            self._period_ioc_miss += 1
            self.timeout_count += 1
            self._log_ioc_miss(miss_reason="TIMEOUT", detect_method="타임아웃")
            self._log_event("TIMEOUT_A", "DT" if is_dt else "DC", self.entry_premium, 0.0,
                            self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate,
                            elapsed * 1000, "Leg A 타임아웃")
            self._last_leg_a_attempt_ts = time.time() + self.cfg.ioc_miss_cooldown_sec
            self._reset()
            return

        # ── 체결 확인: WS 이벤트 우선, REST fallback ─────────────────────
        if self.dry_run:
            filled, exec_qty = True, self.leg_a_btc_qty
            actual_quote_spent = self.leg_a_notional
        else:
            filled, exec_qty, actual_quote_spent = False, 0.0, 0.0
            # 1) User Data Stream 이벤트 체크 (0ms)
            ws_ev = (
                self.price_feed.pop_fill(int(self.leg_a_order_id))
                if self.price_feed else None
            )
            if ws_ev is not None:
                    status_x = ws_ev.get("X", "")
                    exec_qty = float(ws_ev.get("z", 0) or 0.0)
                    actual_quote_spent = float(ws_ev.get("Z", 0) or 0.0) or self.leg_a_notional
                    if status_x in ("EXPIRED", "CANCELED"):
                        # IOC는 부분체결 후 EXPIRED가 가능하다.
                        # executedQty>0이면 체결로 이어서 Leg B를 진행하고,
                        # 완전 미체결일 때만 IOC miss로 집계한다.
                        if exec_qty > 0:
                            filled = True
                            log.info(
                                f"[{self._now_str()}] Leg A IOC 부분체결 후 {status_x}: "
                                f"qty={exec_qty:.8f} {self.cfg.symbol}"
                            )
                        else:
                            log.info(f"[{self._now_str()}] Leg A IOC miss ({status_x}) → IDLE 복귀")
                            self.ioc_miss_count += 1
                            self._period_ioc_miss += 1
                            self._log_ioc_miss(miss_reason=status_x, detect_method="WS이벤트")
                            self._log_event("IOC_MISS", "DT" if is_dt else "DC", self.entry_premium, 0.0,
                                            self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate,
                                            (time.time() - self._leg_a_submit_ts) * 1000, status_x)
                            self._last_leg_a_attempt_ts = time.time() + self.cfg.ioc_miss_cooldown_sec
                            self._reset()
                            return
                    else:
                        filled = status_x == "FILLED"
            else:
                # 2) REST fallback — 최소 간격으로 throttle
                now = time.time()
                if now - self._last_rest_check_ts < _REST_ORDER_CHECK_INTERVAL:
                    return   # 아직 체크 간격 안 됨
                self._last_rest_check_ts = now
                status = self.binance.get_order_status(symbol, self.leg_a_order_id)
                order_status_str  = status.get("status", "")
                filled            = order_status_str == "FILLED"
                exec_qty          = float(status.get("executedQty", 0.0))
                actual_quote_spent = float(status.get("cummulativeQuoteQty", 0.0)) or self.leg_a_notional
                # IOC 주문은 거래소가 즉시 EXPIRED/CANCELED 처리.
                # WS 이벤트 누락 시 REST 폴백이 이를 감지하지 못해 120초 타임아웃까지
                # 대기하는 문제 수정: EXPIRED/CANCELED 즉시 IOC miss로 처리.
                if order_status_str in ("EXPIRED", "CANCELED") and not filled:
                    if exec_qty > 0:
                        # 부분체결 → Leg B 진행
                        filled = True
                        log.info(
                            f"[{self._now_str()}] Leg A IOC 부분체결 후 {order_status_str} (REST): "
                            f"qty={exec_qty:.8f} {self.cfg.symbol}"
                        )
                    else:
                        log.info(
                            f"[{self._now_str()}] Leg A IOC miss ({order_status_str}) REST 감지 → IDLE 복귀"
                        )
                        self.ioc_miss_count += 1
                        self._period_ioc_miss += 1
                        self._log_ioc_miss(miss_reason=order_status_str, detect_method="REST폴백")
                        self._log_event(
                            "IOC_MISS", "DT" if is_dt else "DC", self.entry_premium, 0.0,
                            self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate,
                            (time.time() - self._leg_a_submit_ts) * 1000, order_status_str,
                        )
                        self._last_leg_a_attempt_ts = time.time() + self.cfg.ioc_miss_cooldown_sec
                        self._reset()
                        return

        if filled and exec_qty > 0:
            self.leg_a_btc_qty   = exec_qty
            self.leg_a_fill_time = time.time()
            self._last_rest_check_ts = 0.0   # reset for Leg B
            # leg_a_notional 및 leg_a_price 를 실제 체결 기준으로 보정
            # leg_a_price 는 원래 ask × (1+buffer) 의 지정가인데, IOC 는 실제 ask 에서
            # 체결되므로 entry_premium / abort 계산에 쓰이는 분모가 부풀려짐.
            # → cummulativeQuoteQty / executedQty = 실제 평균 체결가로 교체.
            if not self.dry_run and actual_quote_spent > 0:
                log.debug(
                    f"[{self._now_str()}] Leg A notional 보정: "
                    f"{self.leg_a_notional:.4f} → {actual_quote_spent:.4f}"
                )
                self.leg_a_notional = actual_quote_spent
                if exec_qty > 0:
                    actual_avg_price = actual_quote_spent / exec_qty
                    log.debug(
                        f"[{self._now_str()}] Leg A price 보정: "
                        f"{self.leg_a_price:.6f} → {actual_avg_price:.6f}"
                    )
                    self.leg_a_price = actual_avg_price
            log.info(f"[{self._now_str()}] Leg A 체결: {exec_qty:.8f} {self.cfg.symbol}  notional={self.leg_a_notional:.4f}")
            if is_dt:
                self.usdc = max(self.usdc - self.leg_a_notional, 0.0)
            else:
                self.usdt = max(self.usdt - self.leg_a_notional, 0.0)
            self._place_leg_b()

    # ── Leg B 진입 ────────────────────────────────────────────────────────

    def _place_leg_b(self) -> None:
        mid_ut, mid_uc = self._get_mids()
        is_dt = self.current_action == ActionV2.TRADE_DT
        symbol = self.cfg.binance_symbol_usdt if is_dt else self.cfg.binance_symbol_usdc
        price_base = mid_ut if is_dt else mid_uc

        # ── Leg B 진입 전 수익성 재확인 ───────────────────────────────────
        # Leg A 지정가(= 시그널 시점 mid)와 현재 Leg B mid 비교
        # abort_threshold 미만이면 즉시 시장가 청산
        current_premium = (price_base - self.leg_a_price) / self.leg_a_price
        fee_stack = 2.0 * (self._eff_fee() + self.cfg.slippage_rate)
        abort_threshold = (
            self.cfg.leg_b_abort_threshold_rate
            if self.cfg.leg_b_abort_threshold_rate is not None
            else fee_stack
        )

        if current_premium < abort_threshold:
            total_elapsed = time.time() - self.leg_a_fill_time
            # 진입 시그널은 (USDT-USDC)/USDC 기준 → DC는 음수
            # current_premium은 (Leg-B-sell - Leg-A-buy)/Leg-A-buy 기준 → 양수=수익, 음수=손실
            direction = "DT" if is_dt else "DC"
            entry_margin = abs(self.entry_premium)  # 진입 시 예상 수익률
            status = "스프레드 역전" if current_premium < 0 else "수익성 소멸"
            log.warning(
                f"[{self._now_str()}] ⚡ Leg B 조기 청산: {status} "
                f"[{direction}] 진입예상={entry_margin:.4%} 현재수익성={current_premium:.4%} "
                f"abort임계={abort_threshold:.4%} → 시장가 청산"
            )
            self._notify(
                f"⚡ <b>Leg B 조기 청산</b> ({status})\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"방향: {direction}  |  진입예상수익: +{entry_margin:.4%}\n"
                f"현재 수익성: {current_premium:+.4%}  (abort 임계: {abort_threshold:+.4%})\n"
                f"{'※ 스프레드 완전 역전 — 손절 청산' if current_premium < 0 else '※ 수익 마진 부족 — 예방 청산'}\n"
                f"{self.leg_a_btc_qty:.8f} {self.cfg.symbol} 시장가 청산"
            )
            if not self.dry_run:
                try:
                    self.binance.place_market_order(symbol, "SELL", self.leg_a_btc_qty)
                except Exception as e:
                    log.error(f"Leg B 조기 청산 실패: {e}")
            self.abort_count += 1
            self._hour_abort_count += 1
            self._period_abort_count += 1
            self.leg_b_price = price_base  # 참고용
            self._log_event("ABORT", direction, abs(self.entry_premium), current_premium,
                            abort_threshold, (time.time() - self.leg_a_fill_time) * 1000, status)
            # 실제 손실 계산: Leg A taker 수수료 + 시장가 청산 taker 수수료 + 슬리피지
            abort_sell_notional = self.leg_a_btc_qty * price_base
            abort_received = abort_sell_notional * (1.0 - self.cfg.emergency_taker_fee_rate - self.cfg.emergency_slippage_rate)
            abort_pnl = (
                abort_received
                - self.leg_a_notional
                - self.leg_a_notional * self.cfg.fee_rate
            )
            self.total_pnl_usd += abort_pnl
            self._hour_pnl += abort_pnl
            self._finalize_trade(
                received=abort_received,
                pnl=abort_pnl,
                elapsed_sec=total_elapsed,
                note=f"ABORT_LEG_B_PREMIUM_GONE(abort_at={current_premium:.4%})",
            )
            if is_dt:
                self.usdt = max(self.usdt + abort_received, 0.0)
            else:
                self.usdc = max(self.usdc + abort_received, 0.0)
            self.sync_balances()
            self._reset()
            return
        # ── 수익 가능 → 지정가 Leg B 진입 ───────────────────────────────
        # Leg B 가격 = max(현재시세, 손익분기 최소가격)
        # 현재시세가 이미 수익권이면 빠르게 체결, 아니면 min_profitable_price에 걸고 회복 대기
        full_fee_stack = self.cfg.fee_rate + self._eff_fee() + 2.0 * self.cfg.slippage_rate
        min_profitable_price = self.leg_a_price * (1.0 + full_fee_stack)
        # BUG-25: normalize_price(floor)를 min_profitable_price에 적용하면 틱 단위로 내려가
        # 실제 손익분기 이하가 됨. min_profitable_price는 ceil 적용, 시장가는 floor 적용.
        sell_price = max(
            self.binance.normalize_price(symbol, price_base),
            self.binance.normalize_price_ceil(symbol, min_profitable_price),
        )

        log.info(
            f"[{self._now_str()}] Leg B: SELL {self.leg_a_btc_qty:.8f} {self.cfg.symbol} "
            f"@ {sell_price:.2f}  symbol={symbol}  "
            f"premium={current_premium:.4%} (>= fee_stack {fee_stack:.4%})"
        )

        if self.dry_run:
            self.leg_b_order_id = "DRY"
        else:
            try:
                if self.ws_order_client and self.ws_order_client.is_ready():
                    resp = self.ws_order_client.place_limit_order(
                        symbol, "SELL", self.leg_a_btc_qty, sell_price, time_in_force="GTC"
                    )
                else:
                    resp = self.binance.place_limit_order(
                        symbol, "SELL", self.leg_a_btc_qty, sell_price
                    )
                self.leg_b_order_id = str(resp["orderId"])
            except Exception as e:
                log.error(f"Leg B 주문 실패: {e} — 시장가 조기 청산")
                # Leg B 주문 자체가 실패 → 보유 코인 즉시 시장가 청산
                if not self.dry_run:
                    try:
                        self.binance.place_market_order(symbol, "SELL", self.leg_a_btc_qty)
                    except Exception as e2:
                        log.error(f"Leg B 시장가 청산도 실패: {e2}")
                self.abort_count += 1
                self._hour_abort_count += 1
                self._period_abort_count += 1
                abort_sell_notional2 = self.leg_a_btc_qty * price_base
                abort_received2 = abort_sell_notional2 * (1.0 - self.cfg.emergency_taker_fee_rate - self.cfg.emergency_slippage_rate)
                abort_pnl2 = (
                    abort_received2
                    - self.leg_a_notional
                    - self.leg_a_notional * self.cfg.fee_rate
                )
                self.total_pnl_usd += abort_pnl2
                self._hour_pnl += abort_pnl2
                self._finalize_trade(
                    received=abort_received2,
                    pnl=abort_pnl2,
                    elapsed_sec=time.time() - self.leg_a_fill_time,
                    note="ABORT_LEG_B_ORDER_FAILED",
                )
                if is_dt:
                    self.usdt = max(self.usdt + abort_received2, 0.0)
                else:
                    self.usdc = max(self.usdc + abort_received2, 0.0)
                self.sync_balances()
                self._reset()
                return

        self.leg_b_price = sell_price
        # entry_premium = signal_premium 유지 (재계산하면 Leg B 진입 시점 bid 변동으로 달라짐)
        self.state = TradeState.LEG_B_PENDING
        self.order_time = time.time()

    # ── Leg B 모니터링 ────────────────────────────────────────────────────

    def _leg_b_tick(self) -> None:
        # 타임아웃은 leg_a_fill_time 기준 절대 시간 (reprice로 리셋되지 않음)
        elapsed = time.time() - self.leg_a_fill_time
        is_dt = self.current_action == ActionV2.TRADE_DT
        symbol = self.cfg.binance_symbol_usdt if is_dt else self.cfg.binance_symbol_usdc

        if elapsed >= self.leg_timeout_sec:
            log.warning(
                f"[{self._now_str()}] Leg B 타임아웃 ({elapsed:.1f}s) "
                f"→ 비상 시장가 청산 {self.leg_a_btc_qty:.8f} {self.cfg.symbol}"
            )
            if not self.dry_run:
                try:
                    self.binance.cancel_order(symbol, self.leg_b_order_id)
                except Exception:
                    pass
                try:
                    # [BUG-7] Leg B GTC 부분체결 후 잔여 코인만 조회해서 시장가 청산.
                    # leg_a_btc_qty 전체로 팔면 부분체결분이 이미 빠져서 -2010 발생.
                    actual_coin_bal = self.binance.get_spot_balances().get(self.cfg.symbol, 0.0)
                    sell_qty = self.binance.normalize_market_quantity(symbol, actual_coin_bal)
                    if sell_qty > 0:
                        self.binance.place_market_order(symbol, "SELL", sell_qty)
                    else:
                        log.info(f"비상 청산 불필요: {self.cfg.symbol} 잔고 없음 (Leg B 이미 전체 체결)")
                except Exception as e:
                    log.error(f"비상 청산 실패: {e}")

            self.emergency_count += 1
            self._period_emergency_count += 1
            total_elapsed = time.time() - self.leg_a_fill_time
            self._notify(
                f"⚠️ [{self.cfg.symbol}] Leg B 비상 청산 #{self.emergency_count}\n"
                f"{self.leg_a_btc_qty:.8f} {self.cfg.symbol}  elapsed={total_elapsed:.1f}s"
            )
            # 비상청산도 CSV에 기록 — 실제 손실 계산
            em_mids = self._get_mids()
            em_mid = (em_mids[0] + em_mids[1]) / 2
            em_sell_notional = self.leg_a_btc_qty * em_mid
            em_received = em_sell_notional * (1.0 - self.cfg.emergency_taker_fee_rate - self.cfg.emergency_slippage_rate)
            em_pnl = (
                em_received
                - self.leg_a_notional
                - self.leg_a_notional * self.cfg.fee_rate
            )
            self.total_pnl_usd += em_pnl
            self._hour_pnl += em_pnl
            self._finalize_trade(
                received=em_received,
                pnl=em_pnl,
                elapsed_sec=total_elapsed,
                note="EMERGENCY_MARKET_CLOSE",
            )
            if is_dt:
                self.usdt = max(self.usdt + em_received, 0.0)
            else:
                self.usdc = max(self.usdc + em_received, 0.0)
            self.sync_balances()
            self._reset()
            return

        # ── 가격 방어: 5초마다 현재가 체크 (스탑로스 + 리프라이스) ──────
        now = time.time()
        if now - self._last_price_check_ts >= 1.0:
            self._last_price_check_ts = now
            mid_ut, mid_uc = self._get_mids()
            current_sell_mid = mid_ut if is_dt else mid_uc
            current_gross = (current_sell_mid - self.leg_a_price) / self.leg_a_price if self.leg_a_price > 0 else 0.0

            # 1) 스탑로스: 진입가 대비 stop_loss_rate 이상 하락 시 즉시 청산
            stop_loss = self.cfg.leg_b_stop_loss_rate
            if stop_loss is not None and current_gross < -stop_loss:
                log.warning(
                    f"[{self._now_str()}] 🛑 Leg B 스탑로스 발동: "
                    f"현재gross={current_gross:.4%} < -{stop_loss:.4%} → 즉시 시장가 청산"
                )
                if not self.dry_run:
                    try:
                        self.binance.cancel_order(symbol, self.leg_b_order_id)
                    except Exception:
                        pass
                    try:
                        actual_coin_bal = self.binance.get_spot_balances().get(self.cfg.symbol, 0.0)
                        sell_qty = self.binance.normalize_market_quantity(symbol, actual_coin_bal)
                        if sell_qty > 0:
                            self.binance.place_market_order(symbol, "SELL", sell_qty)
                    except Exception as e:
                        log.error(f"스탑로스 청산 실패: {e}")
                self.emergency_count += 1
                self._period_emergency_count += 1
                total_elapsed = time.time() - self.leg_a_fill_time
                em_sell_notional = self.leg_a_btc_qty * current_sell_mid
                em_received = em_sell_notional * (1.0 - self.cfg.emergency_taker_fee_rate - self.cfg.emergency_slippage_rate)
                em_pnl = em_received - self.leg_a_notional - self.leg_a_notional * self.cfg.fee_rate
                self.total_pnl_usd += em_pnl
                self._hour_pnl += em_pnl
                self._notify(
                    f"🛑 <b>Leg B 스탑로스</b> [{self.cfg.symbol}]\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"진입가: {self.leg_a_price:.6f}  현재가: {current_sell_mid:.6f}\n"
                    f"하락률: {current_gross:.4%}  (임계: -{stop_loss:.4%})\n"
                    f"예상손실: {em_pnl:+.4f} USDT"
                )
                self._finalize_trade(received=em_received, pnl=em_pnl,
                                     elapsed_sec=total_elapsed, note=f"STOP_LOSS(drop={current_gross:.4%})")
                if is_dt:
                    self.usdt = max(self.usdt + em_received, 0.0)
                else:
                    self.usdc = max(self.usdc + em_received, 0.0)
                self.sync_balances()
                self._reset()
                return

            # 2) 리프라이스
            # - T < 30s: 수익권 내에서만 reprice (min_profitable_price 하한 유지, 회복 대기)
            # - T >= 30s: 하한 제거, 시장가 추종 → GTC maker 체결로 비상청산 taker fee 절감
            if self.cfg.leg_b_reprice_on_drop and not self.dry_run:
                entry_threshold = self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate
                _late_reprice_mode = elapsed >= 30.0  # elapsed = leg_a_fill_time 기준 절대 경과

                if current_sell_mid < self.leg_b_price * 0.9999:
                    _full_fee_stack = self.cfg.fee_rate + self._eff_fee() + 2.0 * self.cfg.slippage_rate
                    _min_profitable = self.leg_a_price * (1.0 + _full_fee_stack)

                    if _late_reprice_mode:
                        # 30초 이후: floor 없이 현재 시장가 그대로 추종
                        new_sell_price = self.binance.normalize_price(symbol, current_sell_mid)
                    elif current_gross > entry_threshold:
                        # 30초 이전: 수익권 내에서만 reprice (BUG-25: ceil 적용)
                        new_sell_price = max(
                            self.binance.normalize_price(symbol, current_sell_mid),
                            self.binance.normalize_price_ceil(symbol, _min_profitable),
                        )
                    else:
                        new_sell_price = None  # 수익권 이탈 상태, 회복 대기

                    if new_sell_price is not None and new_sell_price < self.leg_b_price:
                        mode_tag = "⏬후기" if _late_reprice_mode else "🔄"
                        log.info(
                            f"[{self._now_str()}] {mode_tag} Leg B 리프라이스: "
                            f"{self.leg_b_price:.6f} → {new_sell_price:.6f}  "
                            f"gross={current_gross:.4%}  elapsed={elapsed:.0f}s"
                        )
                        try:
                            self.binance.cancel_order(symbol, self.leg_b_order_id)
                            resp = self.binance.place_limit_order(
                                symbol, "SELL", self.leg_a_btc_qty, new_sell_price
                            )
                            self.leg_b_order_id = str(resp["orderId"])
                            self.leg_b_price = new_sell_price
                            if not _late_reprice_mode:
                                self.order_time = time.time()  # 초기 reprice만 order_time 리셋
                        except Exception as e:
                            log.warning(f"Leg B 리프라이스 실패 (계속 대기): {e}")

        # ── 체결 확인: WS 이벤트 우선, REST fallback ─────────────────────
        if self.dry_run:
            filled   = True
            received = self.leg_a_btc_qty * self.leg_b_price * (1.0 - self._eff_fee())
        else:
            filled, received = False, 0.0
            # 1) User Data Stream 이벤트 체크 (0ms)
            ws_ev = (
                self.price_feed.pop_fill(int(self.leg_b_order_id))
                if self.price_feed else None
            )
            if ws_ev is not None:
                filled   = ws_ev.get("X") == "FILLED"
                received = float(ws_ev.get("Z", 0))
            else:
                # 2) REST fallback — throttle
                now = time.time()
                if now - self._last_rest_check_ts < _REST_ORDER_CHECK_INTERVAL:
                    return
                self._last_rest_check_ts = now
                status   = self.binance.get_order_status(symbol, self.leg_b_order_id)
                filled   = status.get("status") == "FILLED"
                received = float(status.get("cummulativeQuoteQty", 0.0))

        if filled and received > 0:
            if is_dt:
                self.usdt += received
            else:
                self.usdc += received

            # Leg A IOC = taker 수수료, Leg B GTC 지정가 = maker 수수료
            fee_est = self.leg_a_notional * self.cfg.fee_rate + received * self._eff_fee()
            pnl = received - self.leg_a_notional - fee_est
            self.total_pnl_usd += pnl
            self._hour_pnl += pnl
            self.trade_count += 1
            self._hour_trade_count += 1
            if self.global_lock is not None:
                self.global_lock.set_last_trade(
                    self.cfg.symbol,
                    self.current_action.value if self.current_action else "",
                )

            total_elapsed = time.time() - self.leg_a_fill_time
            log.info(
                f"[{self._now_str()}] ✅ #{self.trade_count}  "
                f"pnl={pnl:+.4f}  누적={self.total_pnl_usd:+.4f}  "
                f"USDT={self.usdt:.2f}  USDC={self.usdc:.2f}  "
                f"elapsed={total_elapsed:.1f}s"
            )
            self._finalize_trade(received=received, pnl=pnl, elapsed_sec=total_elapsed)

            if self.trade_count % 10 == 0:
                self.sync_balances()

            if self.trade_count % 20 == 0:
                equity = self.usdt + self.usdc
                self._notify(
                    f"📊 <b>v1.28 중간 현황 [{self.cfg.symbol}]</b> (거래 {self.trade_count}회 달성)\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"💰 누적 순수익: {self._usd_krw(self.total_pnl_usd)}  <i>(BNB 수수료 포함)</i>\n"
                    f"💼 잔고\n"
                    f"  USDT: {self.usdt:,.2f}  USDC: {self.usdc:,.2f}\n"
                    f"  BNB: {self.bnb:.6f}\n"
                    f"  총자산: {self._eq_krw(equity)}\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"⚠️ 조기청산: {self.abort_count}회  비상청산: {self.emergency_count}회  IOC miss: {self.ioc_miss_count}회"
                )
                self._send_csv(
                    f"📎 거래기록 (누적 {self.trade_count}건)\n"
                    f"누적 순수익: {self.total_pnl_usd:+.4f} USDT (BNB 수수료 포함)"
                )

            self._reset()

    # ── Speculative 모니터링 ─────────────────────────────────────────────

    def _spec_tick(self) -> None:
        """SPEC_PENDING 상태: Leg A/B 이벤트 동시 모니터링."""
        elapsed = time.time() - self.order_time
        is_dt = self.current_action == ActionV2.TRADE_DT
        sym_a = self.cfg.binance_symbol_usdc if is_dt else self.cfg.binance_symbol_usdt
        sym_b = self.cfg.binance_symbol_usdt if is_dt else self.cfg.binance_symbol_usdc

        if not self.dry_run:
            # ── Leg A 이벤트 확인 ────────────────────────────────────────
            if not self._spec_leg_a_filled and self.leg_a_order_id:
                ws_a = (
                    self.price_feed.pop_fill(int(self.leg_a_order_id))
                    if self.price_feed else None
                )
                # UDS 없을 때 REST 폴링 (0.5s 간격) — listenToken 방식 전환 전 fallback
                if ws_a is None:
                    now = time.time()
                    if now - self._last_rest_check_ts >= _REST_ORDER_CHECK_INTERVAL:
                        self._last_rest_check_ts = now
                        try:
                            st = self.binance.get_order_status(sym_a, self.leg_a_order_id)
                            st_status = st.get("status", "")
                            if st_status in ("FILLED", "EXPIRED", "CANCELED"):
                                exec_qty = float(st.get("executedQty", 0.0))
                                actual_notional = float(st.get("cummulativeQuoteQty", 0.0)) or self.leg_a_notional
                                # ws_a 포맷으로 변환하여 아래 WS 처리 로직 재사용
                                ws_a = {"X": st_status, "z": exec_qty, "Z": actual_notional}
                        except Exception as e:
                            log.debug(f"SPEC Leg A REST 폴링 실패 (무시): {e}")
                if ws_a is not None:
                    status_x = ws_a.get("X", "")
                    exec_qty = float(ws_a.get("z", 0) or 0.0)
                    actual_notional = float(ws_a.get("Z", 0) or 0.0) or self.leg_a_notional
                    if status_x in ("EXPIRED", "CANCELED"):
                        # SPEC도 IOC 부분체결 후 EXPIRED가 가능하다.
                        if exec_qty > 0:
                            self.leg_a_btc_qty = exec_qty
                            self.leg_a_notional = actual_notional
                            self.leg_a_fill_time = time.time()
                            if is_dt:
                                self.usdc = max(self.usdc - self.leg_a_notional, 0.0)
                            else:
                                self.usdt = max(self.usdt - self.leg_a_notional, 0.0)
                            self._spec_leg_a_filled = True
                            log.info(
                                f"[{self._now_str()}] SPEC Leg A IOC 부분체결 후 {status_x}: "
                                f"qty={exec_qty:.8f} {self.cfg.symbol}"
                            )
                        else:
                            # Leg A 완전 miss
                            if self._spec_leg_b_filled:
                                # Leg B가 이미 체결된 상태 → 비상 시장가 매수로 복구
                                log.error(
                                    f"[{self._now_str()}] ⚠️ SPEC Leg A {status_x} + Leg B 선체결 "
                                    f"→ {self.leg_a_btc_qty:.8f} {self.cfg.symbol} 시장가 매수 복구"
                                )
                                self._notify(
                                    f"🚨 <b>SPEC 비상복구 [{self.cfg.symbol}]</b>\n"
                                    f"Leg A IOC miss + Leg B 선체결 → 시장가 매수로 복구\n"
                                    f"{self.leg_a_btc_qty:.8f} {self.cfg.symbol}"
                                )
                                try:
                                    self.binance.place_market_order(sym_a, "BUY", self.leg_a_btc_qty)
                                except Exception as e:
                                    log.error(f"SPEC 비상복구 실패: {e}")
                                self.emergency_count += 1
                                self._period_emergency_count += 1
                                self.sync_balances()
                                # Leg A는 미체결 → USDC/USDT 미지출. 복구 매수 비용만 차감.
                                # leg_a_price는 시그널 시점 IOC 가격 → 실제 시장가 근사치로 사용
                                recovery_buy_est = self.leg_a_btc_qty * self.leg_a_price * (
                                    1.0 + self.cfg.emergency_taker_fee_rate + self.cfg.emergency_slippage_rate
                                )
                                recovery_pnl = self._spec_received - recovery_buy_est
                                self.total_pnl_usd += recovery_pnl
                                self.trade_count += 1
                                self._log_trade(
                                    action=(self.current_action.value + "_SPEC_RECOVERY"),
                                    leg_a_symbol=sym_a,
                                    leg_b_symbol=sym_b,
                                    leg_b_price=self.leg_b_price,
                                    received=self._spec_received,
                                    pnl=recovery_pnl,
                                    elapsed_sec=elapsed,
                                    note="LegA_miss_LegB_filled_recovery",
                                )
                            else:
                                # Leg B 미체결 → 정상 취소 후 IDLE
                                log.info(f"[{self._now_str()}] SPEC Leg A IOC miss → Leg B 취소")
                                self._cancel_order_safe(sym_b, self.leg_b_order_id)
                                self.ioc_miss_count += 1
                                self._period_ioc_miss += 1
                                self._log_ioc_miss(miss_reason=f"SPEC_{status_x}", detect_method="WS이벤트")
                                self._log_event("SPEC_MISS", "DT" if is_dt else "DC", self.entry_premium, 0.0,
                                                self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate,
                                                (time.time() - self._leg_a_submit_ts) * 1000, f"SPEC_{status_x}")
                            self._last_leg_a_attempt_ts = time.time() + self.cfg.ioc_miss_cooldown_sec
                            self._reset()
                            return
                    elif status_x == "FILLED":
                        self.leg_a_btc_qty = exec_qty
                        self.leg_a_notional = actual_notional
                        self.leg_a_fill_time = time.time()
                        if is_dt:
                            self.usdc = max(self.usdc - self.leg_a_notional, 0.0)
                        else:
                            self.usdt = max(self.usdt - self.leg_a_notional, 0.0)
                        self._spec_leg_a_filled = True
                        log.info(f"[{self._now_str()}] SPEC Leg A 체결: {exec_qty:.8f} {self.cfg.symbol}")

            # ── SPEC Leg B 가격 재설정 (Leg A 체결 후 최초 1회) ──────────
            # 시그널 시점 고정 bid → Leg A 체결 시점의 현재 bid로 갱신
            # 시장이 이동해도 체결 가능한 가격에 재제출
            if (
                self._spec_leg_a_filled
                and not self._spec_leg_b_filled
                and not self._spec_leg_b_repriced
                and self.leg_b_order_id
            ):
                new_bid = self._get_best_bid(sym_b)
                new_sell_price = self.binance.normalize_price(sym_b, new_bid)
                if new_sell_price != self.leg_b_price:
                    log.info(
                        f"[{self._now_str()}] SPEC Leg B 가격 재설정: "
                        f"{self.leg_b_price:.4f} → {new_sell_price:.4f}"
                    )
                    if not self.dry_run:
                        self._cancel_order_safe(sym_b, self.leg_b_order_id)
                        try:
                            if self.ws_order_client and self.ws_order_client.is_ready():
                                resp_b2 = self.ws_order_client.place_limit_order(
                                    sym_b, "SELL", self.leg_a_btc_qty, new_sell_price, time_in_force="GTC"
                                )
                            else:
                                resp_b2 = self.binance.place_limit_order(
                                    sym_b, "SELL", self.leg_a_btc_qty, new_sell_price
                                )
                            self.leg_b_order_id = str(resp_b2["orderId"])
                            self.leg_b_price = new_sell_price
                            self.order_time = time.time()  # 타임아웃 리셋
                        except Exception as e:
                            log.warning(f"SPEC Leg B 재제출 실패 (기존 주문 유지): {e}")
                self._spec_leg_b_repriced = True

            # ── Leg B 이벤트 확인 ────────────────────────────────────────
            if not self._spec_leg_b_filled and self.leg_b_order_id:
                ws_b = (
                    self.price_feed.pop_fill(int(self.leg_b_order_id))
                    if self.price_feed else None
                )
                # WS 이벤트 없으면 REST 폴링 (0.5s 간격) — Leg B 선체결 포함 모든 케이스 대응
                if ws_b is None:
                    now = time.time()
                    if now - self._last_rest_check_ts >= _REST_ORDER_CHECK_INTERVAL:
                        self._last_rest_check_ts = now
                        try:
                            st = self.binance.get_order_status(sym_b, self.leg_b_order_id)
                            st_status = st.get("status", "")
                            if st_status in ("FILLED", "EXPIRED", "CANCELED"):
                                ws_b = {
                                    "X": st_status,
                                    "z": float(st.get("executedQty", 0.0)),
                                    "Z": float(st.get("cummulativeQuoteQty", 0.0)),
                                }
                                log.debug(
                                    f"[{self._now_str()}] SPEC Leg B REST 폴링: {st_status}"
                                )
                        except Exception as e:
                            log.debug(f"SPEC Leg B REST 폴링 실패 (무시): {e}")
                if ws_b is not None and ws_b.get("X") == "FILLED":
                    self._spec_received = float(ws_b.get("Z", 0))
                    self._spec_leg_b_filled = True
                    log.info(f"[{self._now_str()}] SPEC Leg B 체결: received={self._spec_received:.4f}")

            # ── Leg B 먼저 체결됐는데 Leg A가 miss된 경우 (위험) ─────────
            if self._spec_leg_b_filled and not self._spec_leg_a_filled:
                # Leg A 타임아웃 확인
                if elapsed >= self.leg_timeout_sec:
                    log.error(
                        f"[{self._now_str()}] ⚠️ SPEC 위험: Leg B 체결됐으나 Leg A 미체결 "
                        f"→ {self.leg_a_btc_qty:.8f} {self.cfg.symbol} 시장가 매수 복구"
                    )
                    self._notify(
                        f"🚨 <b>SPEC 비상복구 [{self.cfg.symbol}]</b>\n"
                        f"Leg B 체결 후 Leg A 미체결 → 시장가 매수로 복구\n"
                        f"{self.leg_a_btc_qty:.8f} {self.cfg.symbol}"
                    )
                    try:
                        self.binance.place_market_order(sym_a, "BUY", self.leg_a_btc_qty)
                    except Exception as e:
                        log.error(f"SPEC 비상복구 실패: {e}")
                    self.emergency_count += 1
                    self._period_emergency_count += 1
                    self.sync_balances()
                    # CSV 기록: Leg B 체결 + Leg A 타임아웃 → 시장가 매수 복구
                    # Leg A는 미체결이므로 leg_a_notional 미지출 → 복구 매수 비용만 차감
                    recovery_buy_est = self.leg_a_btc_qty * self.leg_a_price * (
                        1.0 + self.cfg.emergency_taker_fee_rate + self.cfg.emergency_slippage_rate
                    )
                    recovery_pnl = self._spec_received - recovery_buy_est
                    self.total_pnl_usd += recovery_pnl
                    self.trade_count += 1
                    self._log_trade(
                        action=(self.current_action.value + "_SPEC_RECOVERY"),
                        leg_a_symbol=sym_a,
                        leg_b_symbol=sym_b,
                        leg_b_price=self.leg_b_price,
                        received=self._spec_received,
                        pnl=recovery_pnl,
                        elapsed_sec=elapsed,
                        note=f"LegB_filled_LegA_miss_recovery",
                    )
                    self._reset()
                    return
        else:
            # dry_run
            self._spec_leg_a_filled = True
            self._spec_received = self.leg_a_btc_qty * self.leg_b_price * (1.0 - self._eff_fee())
            self._spec_leg_b_filled = True

        # ── 두 레그 모두 체결 → 완료 ─────────────────────────────────────
        if self._spec_leg_a_filled and self._spec_leg_b_filled:
            received = self._spec_received
            if is_dt:
                self.usdt += received
            else:
                self.usdc += received
            fee_est = self.leg_a_notional * self.cfg.fee_rate + received * self._eff_fee()
            pnl = received - self.leg_a_notional - fee_est
            self.total_pnl_usd += pnl
            self._hour_pnl += pnl
            self.trade_count += 1
            self._hour_trade_count += 1
            total_elapsed = time.time() - self.leg_a_fill_time
            log.info(
                f"[{self._now_str()}] ✅ SPEC #{self.trade_count}  "
                f"pnl={pnl:+.4f}  누적={self.total_pnl_usd:+.4f}  elapsed={total_elapsed:.1f}s"
            )
            self._finalize_trade(received=received, pnl=pnl, elapsed_sec=total_elapsed, note="SPEC")
            if self.trade_count % 10 == 0:
                self.sync_balances()
            self._reset()
            return

        # ── 타임아웃 ─────────────────────────────────────────────────────
        if elapsed >= self.leg_timeout_sec:
            if not self._spec_leg_a_filled:
                # REST로 Leg A 실제 체결 여부 확인 (WS 이벤트 누락 대비 — 잔고 desync 방지)
                if not self.dry_run and self.leg_a_order_id:
                    try:
                        status_a = self.binance.get_order_status(sym_a, self.leg_a_order_id)
                        if status_a.get("status") == "FILLED":
                            exec_qty = float(status_a.get("executedQty", 0.0))
                            actual_notional = float(status_a.get("cummulativeQuoteQty", 0.0)) or self.leg_a_notional
                            self.leg_a_btc_qty = exec_qty
                            self.leg_a_notional = actual_notional
                            self.leg_a_fill_time = time.time()
                            if is_dt:
                                self.usdc = max(self.usdc - self.leg_a_notional, 0.0)
                            else:
                                self.usdt = max(self.usdt - self.leg_a_notional, 0.0)
                            self._spec_leg_a_filled = True
                            log.info(
                                f"[{self._now_str()}] SPEC Leg A 타임아웃 REST 확인: 실제 체결됨 "
                                f"{exec_qty:.8f} {self.cfg.symbol} — 계속 모니터링"
                            )
                            self.order_time = time.time()  # 타임아웃 리셋하여 Leg B 계속 모니터링
                            return
                    except Exception as e:
                        log.warning(f"SPEC Leg A 타임아웃 REST 확인 실패 (무시): {e}")
                # Leg A 미체결 확정 → Leg B 취소 → IDLE
                log.info(f"[{self._now_str()}] SPEC 타임아웃: Leg A 미체결 → Leg B 취소")
                self._cancel_order_safe(sym_b, self.leg_b_order_id)
                self.ioc_miss_count += 1
                self._period_ioc_miss += 1
                self._log_ioc_miss(miss_reason="SPEC_TIMEOUT", detect_method="타임아웃")
                self._log_event("SPEC_TIMEOUT", "DT" if is_dt else "DC", self.entry_premium, 0.0,
                                self.cfg.dt_entry_threshold_rate if is_dt else self.cfg.dc_entry_threshold_rate,
                                elapsed * 1000, "SPEC Leg A 타임아웃")
                self._last_leg_a_attempt_ts = time.time() + self.cfg.ioc_miss_cooldown_sec
                self._reset()
            else:
                # Leg A 체결됐으나 Leg B 타임아웃 → REST로 실제 체결 여부 확인 (UDS 410 대비)
                if not self.dry_run and self.leg_b_order_id:
                    try:
                        status_b = self.binance.get_order_status(sym_b, self.leg_b_order_id)
                        if status_b.get("status") == "FILLED":
                            exec_qty_b = float(status_b.get("executedQty", 0.0))
                            self._spec_received = float(status_b.get("cummulativeQuoteQty", 0.0)) or (
                                exec_qty_b * self.leg_b_price
                            )
                            self._spec_leg_b_filled = True
                            log.info(
                                f"[{self._now_str()}] SPEC Leg B 타임아웃 REST 확인: 실제 체결됨 "
                                f"{exec_qty_b:.8f} {self.cfg.symbol} — 비상청산 불필요"
                            )
                            return  # 다음 tick에서 두 레그 모두 체결 → 정상 완료 처리
                    except Exception as e:
                        log.warning(f"SPEC Leg B 타임아웃 REST 확인 실패 (무시): {e}")
                # Leg B 미체결 확정 → 비상 청산
                log.warning(f"[{self._now_str()}] SPEC Leg B 타임아웃 → 비상 시장가 청산")
                self._cancel_order_safe(sym_b, self.leg_b_order_id)
                try:
                    self.binance.place_market_order(sym_b, "SELL", self.leg_a_btc_qty)
                except Exception as e:
                    log.error(f"SPEC 비상 청산 실패: {e}")
                self.emergency_count += 1
                self._period_emergency_count += 1
                em_mid = self._get_mids()[0 if is_dt else 1]
                em_received = self.leg_a_btc_qty * em_mid * (1.0 - self.cfg.emergency_taker_fee_rate - self.cfg.emergency_slippage_rate)
                em_pnl = (
                    em_received
                    - self.leg_a_notional - self.leg_a_notional * self.cfg.fee_rate
                )
                self.total_pnl_usd += em_pnl
                self._hour_pnl += em_pnl
                self._finalize_trade(
                    received=em_received,
                    pnl=em_pnl,
                    elapsed_sec=elapsed,
                    note="SPEC_EMERGENCY_CLOSE",
                )
                if is_dt:
                    self.usdt = max(self.usdt + em_received, 0.0)
                else:
                    self.usdc = max(self.usdc + em_received, 0.0)
                self.sync_balances()
                self._reset()

    def _finalize_trade(
        self,
        received: float,
        pnl: float,
        elapsed_sec: float,
        note: str = "",
    ) -> None:
        is_dt = self.current_action == ActionV2.TRADE_DT
        self._log_trade(
            action=self.current_action.value + ("" if not note else f"_{note}"),
            leg_a_symbol=self.cfg.binance_symbol_usdc if is_dt else self.cfg.binance_symbol_usdt,
            leg_b_symbol=self.cfg.binance_symbol_usdt if is_dt else self.cfg.binance_symbol_usdc,
            leg_b_price=self.leg_b_price,
            received=received,
            pnl=pnl,
            elapsed_sec=elapsed_sec,
            note=note,
        )

    def _reset(self) -> None:
        self.state = TradeState.IDLE
        self.current_action = ActionV2.HOLD
        self.leg_a_order_id = ""
        self.leg_b_order_id = ""
        self.leg_a_btc_qty = 0.0
        self.leg_a_notional = 0.0
        self.leg_a_price = 0.0
        self.leg_b_price = 0.0
        self.entry_premium = 0.0
        self.signal_premium = 0.0
        self._spec_leg_a_filled = False
        self._spec_leg_b_filled = False
        self._spec_received = 0.0
        self._spec_leg_b_repriced = False
        # 통합 풀 모드: 거래 완료/취소 시 글로벌 락 해제
        if self.global_lock is not None:
            self.global_lock.release(self.cfg.symbol)
